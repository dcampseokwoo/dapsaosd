"""엑셀 반영 — openpyxl로 원본 서식 보존, G열 값만 변경.

- 변경 로그: `스테이지 업데이트(26.07)` 시트에 append (반영 행은 노란 배경)
- 이력: `업데이트 내역` 시트 마지막에 1행 append
- 저장: 원본 덮어쓰지 않고 `{원본명}_updated_{YYMMDD}.xlsx`
"""
from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string

import config

log = logging.getLogger(__name__)

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def _build_row_maps(ws):
    """사업자번호→행, 국문 회사명→행 (행 이동/정렬 대비 재탐색용)."""
    bz_col = column_index_from_string(config.COL_BIZNO)
    nm_col = column_index_from_string(config.COL_NAME_KR)
    bizno_map, name_map = {}, {}
    for i in range(config.DATA_START_ROW, ws.max_row + 1):
        bz = str(ws.cell(row=i, column=bz_col).value or "").strip()
        nm = str(ws.cell(row=i, column=nm_col).value or "").strip()
        if bz and bz not in bizno_map:
            bizno_map[bz] = i
        if nm and nm not in name_map:
            name_map[nm] = i
    return bizno_map, name_map


def _resolve_row(ws, rec, bizno_map, name_map):
    """기록된 행의 회사명이 일치하면 그대로, 아니면 사업자번호/이름으로 재탐색."""
    nm_col = column_index_from_string(config.COL_NAME_KR)
    name = str(rec.get("name_kr", "")).strip()
    row = rec.get("row")
    if row and str(ws.cell(row=row, column=nm_col).value or "").strip() == name:
        return row
    bz = str(rec.get("bizno", "")).strip()
    if bz and bz in bizno_map:
        return bizno_map[bz]
    if name and name in name_map:
        return name_map[name]
    return None


def apply_results(results: list[dict], group_desc: str,
                  excel_path: Path | None = None) -> Path:
    """체크포인트 결과 목록을 엑셀에 반영하고 새 파일 경로를 반환.

    results 항목 키: row, bizno, name_kr, old_stage, new_stage, status,
                     applied, confidence, evidence, source_url, note
    """
    excel_path = excel_path or config.EXCEL_PATH
    wb = load_workbook(excel_path)
    ws = wb[config.SHEET_ALL]
    bizno_map, name_map = _build_row_maps(ws)

    applied_count, mismatch = 0, 0
    log_rows = []
    outcomes = []
    for original in results:
        r = dict(original)
        apply_decision = r.get("apply_decision", r.get("applied", False))
        r["apply_decision"] = bool(apply_decision)
        r["excel_applied"] = False
        r["apply_error"] = ""
        r["excel_row"] = ""
        if apply_decision:
            row = _resolve_row(ws, r, bizno_map, name_map)
            if row is None:
                mismatch += 1
                r["excel_applied"] = False
                r["apply_error"] = "행 불일치 — 회사를 찾지 못함"
                r["status"] = "미반영(행 불일치 — 회사를 찾지 못함)"
                log.warning("[row %s] %s — 엑셀에서 회사를 찾지 못해 반영 생략",
                            r.get("row"), r.get("name_kr"))
            else:
                if row != r.get("row"):
                    log.info("[%s] 행 이동 감지: %s → %s", r.get("name_kr"), r.get("row"), row)
                    r["row"] = row
                ws[f"{config.COL_STAGE}{row}"] = r["new_stage"]
                r["excel_applied"] = True
                r["excel_row"] = row
                applied_count += 1
        log_rows.append(r)
        outcomes.append((original, r))

    _append_update_log(wb, log_rows)
    _append_history(wb, group_desc, len(results), applied_count)

    stamp = dt.date.today().strftime("%y%m%d")
    out_path = excel_path.with_name(f"{excel_path.stem}_updated_{stamp}.xlsx")
    try:
        wb.save(out_path)
    except PermissionError:
        # 파일이 엑셀에서 열려 있는 경우 — 다른 이름으로 저장
        alt = excel_path.with_name(
            f"{excel_path.stem}_updated_{dt.datetime.now().strftime('%y%m%d_%H%M%S')}.xlsx")
        log.warning("%s 저장 불가(파일 열림) — %s 로 저장합니다", out_path.name, alt.name)
        wb.save(alt)
        out_path = alt
    log.info("저장 완료: %s (조사 %d건, 반영 %d건, 행 불일치 %d건)",
             out_path, len(results), applied_count, mismatch)
    # 파일 저장 성공 후에만 실제 반영 결과를 호출자 레코드에 전달한다.
    for original, outcome in outcomes:
        original["apply_decision"] = outcome["apply_decision"]
        original["excel_applied"] = outcome["excel_applied"]
        original["apply_error"] = outcome["apply_error"]
        original["excel_row"] = outcome["excel_row"]
        if outcome["apply_error"]:
            original["status"] = outcome["status"]
    return out_path


def _append_update_log(wb, results: list[dict]):
    if config.SHEET_UPDATE_LOG in wb.sheetnames:
        ws = wb[config.SHEET_UPDATE_LOG]
    else:
        ws = wb.create_sheet(config.SHEET_UPDATE_LOG)
        ws.append(config.UPDATE_LOG_HEADER)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    for r in results:
        ws.append([
            r["row"], r["name_kr"], r.get("old_stage", ""),
            r.get("new_stage", ""), r.get("status", ""),
            r.get("confidence", ""), r.get("evidence", ""),
            r.get("source_url", ""), r.get("note", ""),
        ])
        if r.get("excel_applied"):
            for cell in ws[ws.max_row]:
                cell.fill = YELLOW


def _append_history(wb, group_desc: str, total: int, applied: int):
    if config.SHEET_HISTORY in wb.sheetnames:
        ws = wb[config.SHEET_HISTORY]
    else:
        ws = wb.create_sheet(config.SHEET_HISTORY)
        ws.append(["일자", "대상 그룹", "건수", "출처 설명"])

    ws.append([
        dt.date.today().isoformat(),
        group_desc,
        f"조사 {total}건 / 반영 {applied}건",
        "Gemini Google Search grounding 2단계(스크리닝→정밀 검증) 자동 조사",
    ])


def read_logged_rows(excel_path: Path | None = None) -> set[int]:
    """엑셀 로그 시트에 이미 기록된 행 번호(조사 제외 대상). 시트 없으면 CSV 폴백."""
    excel_path = excel_path or config.EXCEL_PATH
    rows: set[int] = set()
    try:
        wb = load_workbook(excel_path, read_only=True)
        if config.SHEET_UPDATE_LOG in wb.sheetnames:
            ws = wb[config.SHEET_UPDATE_LOG]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                try:
                    rows.add(int(row[0]))
                except (TypeError, ValueError):
                    continue
            wb.close()
            return rows
        wb.close()
    except FileNotFoundError:
        pass

    # 폴백: 업로드된 26.07 로그 CSV
    import csv
    fallback = config.UPDATE_LOG_CSV_FALLBACK
    if fallback.exists():
        with open(fallback, newline="", encoding="utf-8-sig") as f:
            for rec in csv.DictReader(f):
                try:
                    rows.add(int(rec["행"]))
                except (KeyError, TypeError, ValueError):
                    continue
    return rows
