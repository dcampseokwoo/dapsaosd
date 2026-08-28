"""US FORGED — 산출물 워크북(§6) + 자체 채점·리젝트 감사(§8).

시트: 1) 요약+provenance+자체채점  2) 발송_리스트(이메일 보유)  3) 연락처_확보_필요
      4) 리젝트_감사(무작위 30, §8-⑤)  5) 중복_엔티티(§2)  6) 명시_배제(known_exclusions)
      7) 치료제_배제(therapeutics, v6)  8) 스테이지_미상

출력 원칙(§0): "선발/요건충족" 표현 금지. 모든 통과행에 unverifiable_requirements 병기.
"""
from __future__ import annotations

import random
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from engine import engine_dedup, engine_shortlist, engine_snapshot

OUT = Path(__file__).resolve().parent.parent / "output" / "screening" / "us_forged_shortlist.xlsx"
FONT = "Arial"
HDR = PatternFill("solid", fgColor="1F3864")
HF = Font(name=FONT, bold=True, color="FFFFFF", size=9)
GREEN = PatternFill("solid", fgColor="E2EFDA")
YEL = PatternFill("solid", fgColor="FFF2CC")
GRY = PatternFill("solid", fgColor="EDEDED")
THIN = Side(style="thin", color="BFBFBF")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _c(ws, r, c, v, *, bold=False, fill=None, wrap=False, align="left", size=9):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=FONT, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORD
    if fill:
        cell.fill = fill
    return cell


def _hdr(ws, r, hs, ws_):
    from openpyxl.utils import get_column_letter
    for i, (h, w) in enumerate(zip(hs, ws_), 1):
        c = _c(ws, r, i, h, bold=True, fill=HDR, align="center"); c.font = HF
        ws.column_dimensions[get_column_letter(i)].width = w


def _contact_status(e):
    if e.get("email"):
        return "email"
    if e.get("website"):
        return "website_only"
    return "none"


def build(run_ts: str | None = None) -> Path:
    rows = engine_snapshot.load_rows()
    assessed = engine_shortlist.build(rows)
    prov = engine_snapshot.run_metadata(engine_snapshot.DEFAULT_SNAPSHOT,
                                    run_ts or datetime.now(timezone.utc).isoformat(timespec="seconds"),
                                    rows)
    disp = engine_shortlist.summarize(assessed)
    _TORD = {"T1": 0, "T2": 1, "T3": 2}
    send = sorted([a for a in assessed if a["disposition"] == "send"],
                  key=lambda x: (_TORD.get(x.get("tier"), 9),
                                 x.get("cls_matched_program_field", ""), x["name_ko"]))
    with_email = [a for a in send if a.get("email")]
    no_email = [a for a in send if not a.get("email")]

    wb = Workbook()
    # ---- 시트1 요약 + provenance + 자체채점(§8) ----
    ws = wb.active; ws.title = "요약_채점"; ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, "US FORGED 발송 후보 선별 — 요약·provenance·자체채점", bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    r = 3
    _c(ws, r, 1, "provenance", bold=True, fill=GRY); r += 1
    for k in ("input_snapshot", "input_sha256", "input_rows", "engine_commit", "run_timestamp"):
        _c(ws, r, 1, k); _c(ws, r, 2, str(prov.get(k))); r += 1
    r += 1
    _c(ws, r, 1, "판정 분포(전 엔티티 %d)" % len(assessed), bold=True, fill=GRY); r += 1
    for k, v in disp.items():
        _c(ws, r, 1, k); _c(ws, r, 2, v, align="center"); r += 1
    r += 1
    # 자체채점 지표(§8)
    from engine import engine_golden
    _, gsum = engine_golden.evaluate_all()
    mp = gsum.get("classification_must_pass", {}); mf = gsum.get("classification_must_fail", {})
    ELEVEN = ["Robotics/Automation", "Advanced Manufacturing", "Energy/Climate Tech",
              "Industrial Hardware", "Semiconductor/Advanced Materials", "Sensor/Edge Device",
              "Physical AI", "Healthtech Device", "Manufacturing Process Innovation",
              "Aerospace", "Quantum"]
    ff = Counter(a.get("cls_matched_program_field") for a in send)
    zero = [f for f in ELEVEN if ff.get(f, 0) == 0]
    stage_leak = sum(1 for a in send
                     if a["stage_bucket"] == engine_shortlist.engine_stage.OUT_OF_SCOPE)
    tiers = Counter(a.get("tier") for a in send)
    metrics = [
        ("골든 must_pass 통과율", f"{mp.get('pass')}/{mp.get('total')}"),
        ("골든 must_fail 차단율", f"{mf.get('pass')}/{mf.get('total')}"),
        ("스테이지 이탈 잔류(0이어야)", stage_leak),
        ("11개 분야 중 통과 0", ", ".join(zero) or "없음"),
        ("발송 리스트 합계", len(send)),
        ("  T1 / T2 / T3", f"{tiers.get('T1',0)} / {tiers.get('T2',0)} / {tiers.get('T3',0)}"),
        ("  이메일 보유 / 연락처 필요", f"{len(with_email)} / {len(no_email)}"),
    ]
    _c(ws, r, 1, "자체 채점(§8)", bold=True, fill=GRY); r += 1
    for k, v in metrics:
        _c(ws, r, 1, k); _c(ws, r, 2, str(v), align="center"); r += 1
    r += 1
    _c(ws, r, 1, "⚠ 경고: 투자 스테이지 컬럼('Seed' 값 자체)은 신뢰 불가. 결측 255건 외에 "
       "상장 대기업이 'Seed'로 오기재된 사례가 최소 3건 확인됨 — 휴젤(코스닥, 매출 4,251억)·"
       "올릭스(코스닥 226950, RNAi 신약)·한국비엔씨(코스닥 256840). 이 3건은 명시 배제 처리했으나, "
       "동일 유형이 더 있을 수 있으니 스테이지 기반 판정을 과신하지 말 것 — 최종 적격은 설문으로 확인.",
       bold=True, fill=YEL, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.row_dimensions[r].height = 56; r += 1

    # ---- 시트2/3 발송 후보 + 연락처 필요 ----
    TFILL = {"T1": GREEN, "T2": YEL, "T3": GRY}
    NOTE = ("티어: T1 hardtech·high·플래그없음 / T2 hardtech+consumer_facing|maturity / "
            "T3 unclear|저신뢰. 애매해도 배제 아니라 후순위 발송. ‘요건충족/선발’ 아님. "
            "※ 모든 대상은 Lab-scale 프로토타입·미국 진출 의지·창업자/CTO 기술 차별성이 "
            "DB 미검증 — 설문으로 확인(타겟시장에 ‘미국’ 기재된 소수는 진출 의지 일부 확인됨).")

    def _write_candidates(sheet, data, title, need_website=False):
        w = wb.create_sheet(sheet); w.sheet_view.showGridLines = False
        _c(w, 1, 1, title, bold=True, size=12)
        w.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
        _c(w, 2, 1, NOTE, size=8, wrap=True)
        w.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
        w.row_dimensions[2].height = 40
        cols = ["티어", "국문명", "사업자번호", "분야", "스테이지", "타겟시장", "confidence",
                "대표자 이메일", "Website", "1줄 사업 소개(원문)", "판정근거(evidence 전문)",
                "consumer_facing", "maturity_signal", "상장/대형의심"]
        _hdr(w, 3, cols, [7, 17, 14, 20, 10, 12, 9, 24, 24, 44, 44, 12, 18, 18])
        rr = 4
        for a in data:  # 이미 티어→분야→사명 순 정렬
            f = TFILL.get(a.get("tier"), GREEN)
            _c(w, rr, 1, a.get("tier", ""), fill=f, align="center", bold=True)
            _c(w, rr, 2, a["name_ko"], fill=f)
            _c(w, rr, 3, a.get("biz_no", ""), fill=f, size=8)
            _c(w, rr, 4, a.get("cls_matched_program_field", ""), fill=f)
            _c(w, rr, 5, a.get("stage", ""), fill=f, align="center")
            _c(w, rr, 6, a.get("target") or "미상", fill=f, align="center", size=8)
            _c(w, rr, 7, a.get("cls_confidence", ""), fill=f, align="center", size=8)
            _c(w, rr, 8, a.get("email", ""), fill=f, size=8)
            _c(w, rr, 9, a.get("website", ""), fill=f, size=8)
            _c(w, rr, 10, a.get("desc", ""), fill=f, size=8, wrap=True)
            _c(w, rr, 11, a.get("cls_evidence", "") or "", fill=f, size=8, wrap=True)
            _c(w, rr, 12, "⚠" if a.get("cls_consumer_facing_end_product") else "", fill=f, align="center", size=8)
            _c(w, rr, 13, a.get("cls_maturity_signal") or "", fill=f, size=8, wrap=True)
            _c(w, rr, 14, a.get("established_suspect") or "", fill=f, size=8, wrap=True)
            rr += 1
        w.freeze_panes = "A4"

    _write_candidates("발송_리스트", with_email,
                      f"발송 리스트(이메일 보유) {len(with_email)}개사 — T1→T3 우선순위 정렬")
    _write_candidates("연락처_확보_필요", no_email,
                      f"발송 리스트(이메일 결측) {len(no_email)}개사 — Website 필수, 확보 후 발송 (T1→T3)",
                      need_website=True)

    # ---- 시트5 리젝트 감사(§8-⑤): 무작위 30(sw15/consumer10/notstartup5) ----
    rnd = random.Random(20260821)
    pools = {"software_only": [a for a in assessed if a.get("cls_verdict") == "software_only"],
             "consumer": [a for a in assessed if a.get("cls_verdict") == "consumer"],
             "not_a_startup": [a for a in assessed if a.get("cls_verdict") == "not_a_startup"]}
    picks = []
    for verd, n in (("software_only", 15), ("consumer", 10), ("not_a_startup", 5)):
        lst = pools[verd][:]; rnd.shuffle(lst); picks += lst[:n]
    wa = wb.create_sheet("리젝트_감사"); wa.sheet_view.showGridLines = False
    _c(wa, 1, 1, "리젝트 감사 — 탈락 풀 무작위 30(sw15/consumer10/비스타트업5, 시드고정). "
       "재현율 확인용: 잘못 탈락한 게 없는지 사람이 읽는다(§8-⑤).", bold=True, size=11, wrap=True)
    wa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    wa.row_dimensions[1].height = 30
    _hdr(wa, 3, ["verdict", "국문명", "1줄 소개(desc)", "판정근거(evidence)", "업종(CB)"],
         [14, 18, 48, 40, 20])
    rr = 4
    for a in picks:
        _c(wa, rr, 1, a.get("cls_verdict"), align="center", size=8)
        _c(wa, rr, 2, a["name_ko"], size=8)
        _c(wa, rr, 3, (a.get("desc") or "")[:70], size=8)
        _c(wa, rr, 4, (a.get("cls_evidence") or "")[:55], size=8)
        _c(wa, rr, 5, (a.get("industry") or "")[:20], size=8)
        rr += 1
    wa.freeze_panes = "A4"

    # ---- 시트6 중복 엔티티(§2) ----
    dups = engine_dedup.duplicate_report(rows)
    wd = wb.create_sheet("중복_엔티티"); wd.sheet_view.showGridLines = False
    _c(wd, 1, 1, f"동명 다중행 {len(dups)}건 — 신원 판정(§2). DB 관리자 정정 참고",
       bold=True, size=12)
    wd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _hdr(wd, 3, ["사명", "행수", "엔티티수", "신원", "플래그", "사업자번호들"],
         [20, 8, 10, 16, 24, 40])
    rr = 4
    for d in dups:
        _c(wd, rr, 1, d["name"]); _c(wd, rr, 2, d["rows"], align="center")
        _c(wd, rr, 3, d["entities"], align="center"); _c(wd, rr, 4, d["identity"], size=8)
        _c(wd, rr, 5, d["flags"], size=8); _c(wd, rr, 6, ", ".join(d["biz_nos"]), size=8)
        rr += 1

    # ---- 시트: 명시 배제(known_exclusions) ----
    from engine import engine_exclude
    known = [a for a in assessed if a.get("biz_no") in engine_exclude.KNOWN_EXCLUDED]
    wk = wb.create_sheet("명시_배제"); wk.sheet_view.showGridLines = False
    _c(wk, 1, 1, f"명시 배제 {len(known)}건 — config/global_exclusions.yaml (규칙 추론 아니라 "
       "사업자번호 명시 관리). 무엇을 왜 뺐는지 노출.", bold=True, size=11, wrap=True)
    wk.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4); wk.row_dimensions[1].height = 26
    _hdr(wk, 3, ["국문명", "사업자번호", "DB 스테이지", "배제 사유"], [22, 16, 12, 70])
    rr = 4
    for a in known:
        _c(wk, rr, 1, a["name_ko"]); _c(wk, rr, 2, a.get("biz_no", ""), size=8)
        _c(wk, rr, 3, a.get("stage", ""), align="center", size=8)
        _c(wk, rr, 4, engine_exclude.KNOWN_EXCLUDED[a["biz_no"]], size=8, wrap=True); rr += 1

    # ---- 시트: 치료제·신약 배제(therapeutics, v6) ----
    ther = sorted([a for a in assessed if a["disposition"] == "excluded_therapeutics"],
                  key=lambda x: x["name_ko"])
    wt = wb.create_sheet("치료제_배제"); wt.sheet_view.showGridLines = False
    _c(wt, 1, 1, f"치료제·신약 배제 {len(ther)}건 (v6) — 치료제·신약 후보물질·백신·항체·의약품 "
       "'자체'를 개발/제조하는 기업은 물리적 하드웨어가 아니므로 발송 제외. "
       "단, 진단기기·수술기구·분석장비·약물전달 디바이스·의료용 소재처럼 '기기·소재'를 만드는 "
       "곳은 hardtech로 발송 리스트에 유지함(경계는 배제 안 하고 T3).", bold=True, size=11, wrap=True)
    wt.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4); wt.row_dimensions[1].height = 48
    _hdr(wt, 3, ["국문명", "사업자번호", "1줄 사업 소개(원문)", "판정근거(evidence)"], [22, 16, 60, 50])
    rr = 4
    for a in ther:
        _c(wt, rr, 1, a["name_ko"]); _c(wt, rr, 2, a.get("biz_no", ""), size=8)
        _c(wt, rr, 3, a.get("desc", ""), size=8, wrap=True)
        _c(wt, rr, 4, a.get("cls_evidence", "") or "", size=8, wrap=True); rr += 1
    wt.freeze_panes = "A4"

    # ---- 시트: 스테이지 미상(발송 리스트) — 사용자 직접 검토용 ----
    unk = [a for a in send if a.get("stage") in ("알 수 없음", "", None)]
    unk.sort(key=lambda x: ({"T1": 0, "T2": 1, "T3": 2}.get(x.get("tier"), 9), x["name_ko"]))
    ws2 = wb.create_sheet("스테이지_미상"); ws2.sheet_view.showGridLines = False
    _c(ws2, 1, 1, f"발송 리스트 중 스테이지 미상 {len(unk)}건 — 사용자 직접 훑고 배제 목록 "
       "추가 표시용. (스테이지 컬럼은 오기재 사례 있음: 휴젤=Seed)", bold=True, size=11, wrap=True)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6); ws2.row_dimensions[1].height = 26
    _hdr(ws2, 3, ["티어", "국문명", "사업자번호", "분야", "1줄 사업 소개(원문)", "상장/대형의심"],
         [7, 20, 14, 22, 60, 18])
    rr = 4
    for a in unk:
        f = TFILL.get(a.get("tier"), GRY)
        _c(ws2, rr, 1, a.get("tier", ""), fill=f, align="center", bold=True)
        _c(ws2, rr, 2, a["name_ko"], fill=f); _c(ws2, rr, 3, a.get("biz_no", ""), fill=f, size=8)
        _c(ws2, rr, 4, a.get("cls_matched_program_field", ""), fill=f, size=8)
        _c(ws2, rr, 5, a.get("desc", ""), fill=f, size=8, wrap=True)
        _c(ws2, rr, 6, a.get("established_suspect") or "", fill=f, size=8); rr += 1
    ws2.freeze_panes = "A4"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT
