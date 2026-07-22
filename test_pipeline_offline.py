"""오프라인 스모크 테스트 — Gemini 호출을 모킹해 파이프라인 전 구간 검증.

실행: python test_pipeline_offline.py
API 키/네트워크 불필요. 합성 Startup_DB.xlsx 생성 → 조사 → 엑셀 반영까지 확인.
"""
import json
import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

import config

TMP = Path(tempfile.mkdtemp(prefix="stage_test_"))
config.CHECKPOINT_DIR = TMP / "checkpoints"
config.CHECKPOINT_PATH = config.CHECKPOINT_DIR / "results.jsonl"

import pipeline
from ai import gemini
from collectors import news_search
from extractors import investment_extractor as ext
from validators import stage_validator as sv

config.SEARCH_MODE = "grounding"  # 기본 흐름은 grounding 모킹으로 검증
config.TARGET_PRIORITIES = {1, 2, 3, 4, 5}  # 테스트는 전체 그룹 대상


def make_excel(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = config.SHEET_ALL
    ws["B3"] = "사업자등록번호"; ws["C3"] = "국문 회사명"; ws["G3"] = "투자 스테이지"
    rows = [
        # row, bizno, kr, en, industry, stage, htype, website
        (4, "111-11-11111", "테스트고침", "TestFix", "AI", "Series A", "Type 1", "a.com"),
        (5, "222-22-22222", "테스트유지", "TestKeep", "커머스", "Seed", "디데이", "b.com"),
        (6, "333-33-33333", "테스트종결", "TestDone", "핀테크", "IPO('24)", "Type 1", "c.com"),
        (7, "444-44-44444", "테스트미디엄", "TestMed", "바이오", "알 수 없음", "Type 2", "d.com"),
        (8, "555-55-55555", "테스트로그기록", "TestLogged", "게임", "Seed", "Type 1", "e.com"),
        (9, "666-66-66666", "테스트연도보정", "TestBare", "로봇", "M&A", "Type 3", "f.com"),
    ]
    for r, bz, kr, en, ind, st, ht, web in rows:
        ws[f"B{r}"] = bz; ws[f"C{r}"] = kr; ws[f"D{r}"] = en; ws[f"E{r}"] = ind
        ws[f"G{r}"] = st; ws[f"H{r}"] = ht; ws[f"M{r}"] = web
    # 로그 시트: row 8은 이미 조사 완료 → 제외돼야 함
    log_ws = wb.create_sheet(config.SHEET_UPDATE_LOG)
    log_ws.append(config.UPDATE_LOG_HEADER)
    log_ws.append([8, "테스트로그기록", "Seed", "Seed", "변경 없음", "high", "", "", ""])
    wb.save(path)


SCRIPTED = {
    "테스트고침": [
        '{"verdict": "changed", "new_stage": "Series B", "note": "시리즈B 기사 확인"}',
        '{"new_stage": "Series B", "confidence": "high", "evidence": "2026.5 시리즈B 100억 (플래텀)", "round_quote": "100억 규모 시리즈B 투자 유치", "article_date": "2026-05", "source_url": "https://platum.kr/x", "note": ""}',
    ],
    "테스트유지": ['{"verdict": "unchanged", "new_stage": "", "note": "시드 유지"}'],
    "테스트미디엄": [
        '{"verdict": "changed", "new_stage": "Seed", "note": "시드 기사"}',
        '{"new_stage": "Seed", "confidence": "medium", "evidence": "2024 시드 (와우테일)", "round_quote": "시드 투자를 유치했다", "article_date": "2024-03", "source_url": "https://wowtale.net/y", "note": ""}',
    ],
    "테스트연도보정": [
        '{"verdict": "unchanged", "new_stage": "", "note": "M&A 확인"}',
        '{"new_stage": "M&A(\'23)", "confidence": "high", "evidence": "2023.2 인수 (더벨)", "round_quote": "지분 전량 인수 계약", "article_date": "2023-02", "source_url": "https://thebell.co.kr/z", "note": ""}',
    ],
}


class FakeClient:
    def __init__(self, **kw):
        self.call_count = 0
        self.max_calls = kw.get("max_calls")
        self._queues = {k: list(v) for k, v in SCRIPTED.items()}

    def grounded(self, prompt, model=None, allowed_models=None):
        if self.max_calls is not None and self.call_count >= self.max_calls:
            raise gemini.BudgetExceeded("한도")
        self.call_count += 1
        for name, queue in self._queues.items():
            if name in prompt and queue:
                return gemini.GroundedAnswer(text=queue.pop(0), sources=[])
        return gemini.GroundedAnswer(text='{"verdict": "no_info", "new_stage": "", "note": ""}', sources=[])

    plain = grounded  # RSS 모드도 같은 시나리오로 응답

    def has_available_models(self, allowed_models=None, preferred=None):
        return True


def main():
    excel = TMP / "Startup_DB.xlsx"
    make_excel(excel)
    pipeline.GeminiClient = FakeClient  # Gemini 모킹

    # ---- 일일 한도 감지/모델 전환 단위 확인
    daily_err = ("429 RESOURCE_EXHAUSTED ... 'quotaId': "
                 "'GenerateRequestsPerDayPerProjectPerModel-FreeTier' ... Please retry in 26.02s")
    assert gemini._is_daily_quota(daily_err)
    assert not gemini._is_daily_quota("429 ... PerMinute ...")
    assert gemini._retry_seconds(daily_err) == 26.02
    # ---- 실제 429를 통한 복수 키 전환 + 이전 키 하위 모델 재사용
    import os
    original_client_factory = gemini.genai.Client
    original_keys = os.environ.get("GEMINI_API_KEYS")
    original_delay = config.REQUEST_DELAY_SEC
    calls = []
    class FakeModels:
        def __init__(self, key): self.key = key
        def generate_content(self, model, contents, config):
            calls.append((self.key, model))
            if self.key == "key-1" and model == "top-a":
                raise gemini.genai_errors.ClientError(
                    429, {"error": {"message": "PerDay quota exhausted"}})
            return type("Resp", (), {"text": "ok", "candidates": []})()
    class FakeSDK:
        def __init__(self, key): self.models = FakeModels(key)
    gemini.genai.Client = lambda api_key: FakeSDK(api_key)
    try:
        os.environ["GEMINI_API_KEYS"] = "key-1,key-2"
        config.REQUEST_DELAY_SEC = 0
        multi = gemini.GeminiClient()
        assert multi.plain("x", model="top-a", allowed_models=["top-a"]).text == "ok"
        assert calls == [("key-1", "top-a"), ("key-2", "top-a")]
        # 현재 키는 2번이지만 1번 키의 미소진 하위 모델을 다시 찾는다.
        multi._blocked_by_key[1]["low"] = "404"
        assert multi._pick_slot("low", ["low"]) == (0, "low")
        multi._blocked_by_key[0]["low"] = "daily"
        assert not multi.has_available_models(["low"], "low")
    finally:
        gemini.genai.Client = original_client_factory
        config.REQUEST_DELAY_SEC = original_delay
        if original_keys is None:
            os.environ.pop("GEMINI_API_KEYS", None)
        else:
            os.environ["GEMINI_API_KEYS"] = original_keys

    # ---- validator 단위 확인
    assert sv.normalize_stage("시리즈B") == "Series B"
    assert sv.normalize_stage("IPO(25)") == "IPO('25)"
    assert sv.is_terminal("M&A('24)") and not sv.is_terminal("M&A")
    assert sv.decide("Series A", "Series B", "medium").apply is False
    assert sv.decide("알 수 없음", "Seed", "medium").apply is True
    assert sv.decide("Seed", "Series A", "high", note="2023 폐업").apply is False

    # ---- 라운드명 인용 검사 (환각 차단)
    assert sv.check_round_quote("Series A", "시리즈A 투자 유치", "high")[0] == "high"
    assert sv.check_round_quote("Series A", "", "high")[0] == "low"            # 인용 부재
    assert sv.check_round_quote("Series A", "31억 규모 후속 투자 유치", "high")[0] == "low"  # 라운드명 없음
    assert sv.check_round_quote("Series A", "프리시리즈A 투자 유치", "high")[0] == "low"     # Pre-A를 Series A로 오인
    assert sv.check_round_quote("Seed", "프리시드 라운드", "high")[0] == "low"
    assert sv.check_round_quote("IPO('25)", "코스닥 상장 완료", "high")[0] == "high"
    assert sv.check_round_quote("M&A('24)", "지분 인수", "medium")[0] == "medium"
    assert sv.check_round_quote("알 수 없음", "", "medium")[0] == "medium"     # 검사 제외

    # ---- 네이버 검색 파서 + 결과 병합 단위 확인
    from collectors import naver_search
    assert naver_search.search_news("아무거나") == []  # 키 미설정 → 빈 결과
    items = naver_search.parse_items([{
        "title": "<b>테스트고침</b>, 시리즈B &quot;100억&quot; 유치",
        "originallink": "https://platum.kr/archives/1",
        "link": "https://n.news.naver.com/x",
        "description": "시리즈B 라운드를 <b>마무리</b>했다",
        "pubDate": "Mon, 20 Jul 2026 09:00:00 +0900",
    }])
    assert items[0]["title"] == '테스트고침, 시리즈B "100억" 유치'
    assert items[0]["link"] == "https://platum.kr/archives/1"  # 원문 링크 우선
    assert items[0]["date"] == "2026-07-20" and items[0]["source"] == "platum.kr"
    merged = news_search.merge_results(
        [{"title": "같은 기사", "link": "a", "date": "2025-01-01"}],
        [{"title": "같은기사", "link": "b", "date": "2025-01-01"},   # 공백 차이 → 중복 제거
         {"title": "최신 기사", "link": "c", "date": "2026-06-01"}],
    )
    assert len(merged) == 2 and merged[0]["title"] == "최신 기사"  # 최신순

    # ---- 사이트 검색 파서 단위 확인
    from collectors import site_search
    wp_html = ('<h2 class="entry-title"><a href="https://platum.kr/archives/1234?utm=x">'
               '테스트고침, 시리즈B 투자 유치</a></h2>'
               '<a href="https://platum.kr/category/news">뉴스</a>')
    found = []
    for url, inner in site_search._A_RE.findall(wp_html):
        found.append((url, inner))
    assert any("archives/1234" in u for u, _ in found)
    art_html = ('<meta property="article:published_time" content="2026-05-10T09:00:00+09:00">'
                '<p>테스트고침이 100억 규모 시리즈B 투자를 유치했다.</p>')
    assert site_search._DATE_META_RE.search(art_html).group(1) == "2026-05-10"

    # 조사 범위 필터: Type 1·디데이·Type 2 만이면 Type 3(테스트연도보정) 제외
    config.TARGET_PRIORITIES = {1, 2, 3}
    scoped = pipeline.select_targets(pipeline.load_companies(excel), excel_path=excel)
    assert all(c.priority in (1, 2, 3) for c in scoped)
    assert "테스트연도보정" not in [c.name_kr for c in scoped]
    config.TARGET_PRIORITIES = {1, 2, 3, 4, 5}

    results = pipeline.run(dry_run=False, excel_path=excel)
    by_name = {r["name_kr"]: r for r in results}

    # 제외 규칙: IPO('24) 종결 + 로그 기록 행은 조사 안 함
    assert "테스트종결" not in by_name and "테스트로그기록" not in by_name
    # high → 반영 / unchanged → 미반영 / medium+알수없음 → 반영 / 연도 보정 → 반영
    assert by_name["테스트고침"]["applied"] and by_name["테스트고침"]["new_stage"] == "Series B"
    assert not by_name["테스트유지"]["applied"]
    assert by_name["테스트미디엄"]["applied"]
    assert by_name["테스트연도보정"]["applied"] and by_name["테스트연도보정"]["new_stage"] == "M&A('23)"

    # 엑셀 출력 검증
    out = next(TMP.glob("Startup_DB_updated_*.xlsx"))
    wb = load_workbook(out)
    ws = wb[config.SHEET_ALL]
    assert ws["G4"].value == "Series B"      # high 반영
    assert ws["G5"].value == "Seed"          # 유지
    assert ws["G7"].value == "Seed"          # medium 공란채움
    assert ws["G9"].value == "M&A('23)"      # 연도 보정
    log_ws = wb[config.SHEET_UPDATE_LOG]
    assert log_ws.max_row == 2 + len(results)  # 기존 1행 + 신규 4행
    assert wb[config.SHEET_HISTORY].max_row >= 2

    # 행 불일치 안전장치: 기록된 행이 틀려도 사업자번호로 재탐색해 올바른 셀에 반영
    from updater import excel_updater
    moved = {"row": 999, "bizno": "222-22-22222", "name_kr": "테스트유지",
             "old_stage": "Seed", "new_stage": "Pre-A", "applied": True,
             "status": "반영", "confidence": "high", "evidence": "", "source_url": "", "note": ""}
    ghost = {"row": 4, "bizno": "000-00-00000", "name_kr": "존재하지않는회사",
             "old_stage": "", "new_stage": "Series C", "applied": True,
             "status": "반영", "confidence": "high", "evidence": "", "source_url": "", "note": ""}
    out2 = excel_updater.apply_results([moved, ghost], "행탐색 테스트", excel)
    wb2 = load_workbook(out2)
    assert wb2[config.SHEET_ALL]["G5"].value == "Pre-A"       # 999행 → 5행 재탐색 성공
    assert wb2[config.SHEET_ALL]["G4"].value == "Series A"    # 유령 회사는 반영 안 됨(원본 유지)

    # 체크포인트 재개: 재실행 시 처리분 스킵 → 조사 0건
    results2 = pipeline.run(dry_run=True, excel_path=excel)
    assert results2 == []

    # 체크포인트 파일 형식 확인
    checkpoint = pipeline.load_checkpoint()
    assert len(checkpoint) == 4 and all("verdict" in r for r in checkpoint.values())

    # ---- 검수 리포트: UTF-8-BOM, 반영/확인필요 분리
    applied_path, applied_count, review_path, review_count = pipeline.export_review_reports()
    assert applied_count == 3 and review_count == 0
    assert applied_path.read_bytes().startswith(b"\xef\xbb\xbf")
    assert review_path.read_bytes().startswith(b"\xef\xbb\xbf")
    assert "회사명" in applied_path.read_text(encoding="utf-8-sig")

    # 실제 엑셀 반영 실패 건은 반영 리포트가 아니라 확인필요 리포트에 포함
    failed = dict(checkpoint[4], row=999, bizno="000", name_kr="유령회사",
                  verdict="changed", applied=True, apply_decision=True,
                  excel_applied=False, apply_error="행 불일치")
    pipeline.append_checkpoint(failed)
    _, applied_count2, review_path2, review_count2 = pipeline.export_review_reports()
    assert applied_count2 == 3 and review_count2 == 1, (applied_count2, review_count2)
    assert "유령회사" in review_path2.read_text(encoding="utf-8-sig")
    # 이후 테스트에 영향을 주지 않도록 같은 행의 정상 최신 레코드를 복원
    pipeline.append_checkpoint(checkpoint[4])

    # 손상된 마지막 줄이 있어도 정상 레코드는 복구되고 손상본은 별도 보관
    with open(config.CHECKPOINT_PATH, "a", encoding="utf-8") as f:
        f.write('{"row":')
    recovered = pipeline.load_checkpoint()
    assert len(recovered) == 5
    assert config.CHECKPOINT_PATH.with_name("results.corrupt.jsonl").exists()

    # ---- 제한된 검증 모델 소진 시 체크포인트 없이 건너뛰기
    class VerifyExhaustedClient(FakeClient):
        def has_available_models(self, allowed_models=None, preferred=None):
            return False
        def grounded(self, prompt, model=None, allowed_models=None):
            if allowed_models:
                raise gemini.AllModelsExhausted("상위 모델 소진")
            return super().grounded(prompt, model, allowed_models)
        plain = grounded
    verify_client = VerifyExhaustedClient()
    verify_target = pipeline.Company(
        row=20, bizno="", name_kr="테스트고침", name_en="", industry="AI",
        stage="Series A", htype="Type 1", website="")
    verify_results = []
    pipeline._run_serial(verify_client, [verify_target], {}, verify_results,
                         {"changed": 0, "unchanged": 0, "no_info": 0,
                          "ambiguous": 0, "applied": 0, "verified": 0})
    assert verify_results == [] and 20 not in pipeline.load_checkpoint()

    # ---- RSS 검색 모드: 파서 + 스크리닝/검증 흐름 확인
    rss_xml = b"""<?xml version="1.0"?><rss><channel>
      <item><title>\xed\x85\x8c\xec\x8a\xa4\xed\x8a\xb8\xea\xb3\xa0\xec\xb9\xa8, 100\xec\x96\xb5 \xec\x8b\x9c\xeb\xa6\xac\xec\xa6\x88B \xed\x88\xac\xec\x9e\x90 \xec\x9c\xa0\xec\xb9\x98</title>
      <link>https://platum.kr/test</link><source>\xed\x94\x8c\xeb\x9e\x98\xed\x85\x80</source>
      <pubDate>Wed, 06 May 2026 09:00:00 GMT</pubDate></item>
    </channel></rss>"""
    parsed = news_search.parse_rss(rss_xml)
    assert parsed[0]["date"] == "2026-05-06" and parsed[0]["link"] == "https://platum.kr/test"
    assert "링크:" in news_search.format_block(parsed)

    # THE VC 수집기 파서 단위 확인
    from collectors import thevc_collector
    ddg_html = '<a href="//duckduckgo.com/l/?uddg=https%3A%2F%2Fthevc.kr%2Fxbarx&rut=abc">x</a>'
    assert thevc_collector._DDG_LINK_RE.findall(ddg_html)
    page_html = ('<title>엑스바엑스 | THE VC</title>'
                 '<meta property="og:description" content="최근 투자 Series A">'
                 '<body>본문</body>')
    assert "Series A" in "".join(thevc_collector._META_RE.findall(page_html))

    # ---- 배치 스크리닝 파서 단위 확인
    batch_ans = type("A", (), {"text": '```json\n[{"id":1,"verdict":"changed","new_stage":"Series B","note":"x"},{"id":2,"verdict":"unchanged","new_stage":"","note":""}]\n```'})()
    bm = ext.parse_screening_batch(batch_ans, 2)
    assert bm[1].verdict == "changed" and bm[1].new_stage == "Series B"
    assert bm[2].verdict == "unchanged"
    assert ext.parse_screening_batch(type("A", (), {"text": "쓰레기"})(), 2) == {}  # 실패→빈dict

    config.SEARCH_MODE = "rss"
    config.SCREEN_BATCH_SIZE = 1  # 통합 run 테스트는 순차 경로로 고정
    config.CHECKPOINT_PATH.unlink()  # 재조사 위해 체크포인트 초기화
    news_search.search_news = lambda q, max_items=8, timeout=15: parsed  # 네트워크 모킹
    site_search.collect_articles = lambda name, max_articles=3: []
    thevc_collector.get_block = lambda name: "[THE VC 페이지 https://thevc.kr/test]\n최근 투자 라운드 Series B"
    results_rss = pipeline.run(dry_run=True, excel_path=excel)
    rss_fix = {r["name_kr"]: r for r in results_rss}["테스트고침"]
    assert rss_fix["applied"] and rss_fix["new_stage"] == "Series B"
    # 검색 결과 0건이면 Gemini 호출 없이 no_info 처리되는지
    news_search.search_news = lambda q, max_items=8, timeout=15: []
    config.CHECKPOINT_PATH.unlink()
    results_empty = pipeline.run(dry_run=True, excel_path=excel)
    assert all(r["verdict"] == "no_info" for r in results_empty
               if r["name_kr"] != "테스트연도보정")

    # ---- 배치 스크리닝 경로 통합 확인 (screen_chunk)
    class BatchClient:
        def __init__(self): self.call_count = 0; self.max_calls = None
        def plain(self, prompt, model=None):
            self.call_count += 1
            assert "[배치 스크리닝]" in prompt  # 배치 프롬프트로 호출됐는지
            # 회사 1=changed, 회사 2=unchanged 로 응답
            return gemini.GroundedAnswer(
                text='[{"id":1,"verdict":"changed","new_stage":"Series B","note":""},'
                     '{"id":2,"verdict":"unchanged","new_stage":"","note":""}]', sources=[])
    # 검색: A/B는 결과 있음, C는 0건(→ API 없이 no_info)
    def fake_search(q, max_items=8, timeout=15):
        return [] if "씨회사" in q else [{"title":"기사","link":"u","date":"2026-01-01"}]
    news_search.search_news = fake_search
    naver_search.search_news = lambda q, max_items=8, sort="date": []
    C = pipeline.Company
    chunk = [C(row=10,bizno="",name_kr="에이회사",name_en="",industry="",stage="Series A",htype="",website=""),
             C(row=11,bizno="",name_kr="비회사",name_en="",industry="",stage="Seed",htype="",website=""),
             C(row=12,bizno="",name_kr="씨회사",name_en="",industry="",stage="Seed",htype="",website="")]
    bc = BatchClient()
    batch_cache = {}
    screened = pipeline.screen_chunk(bc, chunk, {}, batch_cache)
    assert bc.call_count == 1  # 3개 회사를 단 1회 호출로 스크리닝
    assert screened[10][0].verdict == "changed" and screened[10][0].new_stage == "Series B"
    assert screened[11][0].verdict == "unchanged"
    assert screened[12][0].verdict == "no_info"  # 검색 0건은 API 없이 처리
    assert batch_cache[10][0]["title"] == "기사"  # 2단계 검증에서 재사용할 검색 캐시

    print("모든 오프라인 테스트 통과 ✔ (grounding + rss + 배치 모드)")
    print(f"  조사 {len(results)}건 / 반영 {sum(r['applied'] for r in results)}건")
    print(f"  출력: {out.name}")
    shutil.rmtree(TMP)


if __name__ == "__main__":
    main()
