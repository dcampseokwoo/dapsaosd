"""파이프라인 오케스트레이션.

흐름:
  1) 로드     — All(전체기업) 시트에서 대상 로드
  2) 제외     — 26.07 로그 기록분(425) + 연도확정 IPO/M&A('YY)
  3) 정렬     — H열 우선순위 (Type 1 → 디데이 → Type 2 → 공란/기타 → Type 3)
  4) 1단계    — 스크리닝 검색 1회 → changed/unchanged/no_info/ambiguous
               (스크리닝_기완료_50개사.csv 는 1단계 스킵, verdict 재사용)
  5) 2단계    — changed/ambiguous만 정밀 검증 (검색 2-3회 교차 확인)
  6) 판정     — stage_validator.decide (보수적 반영 정책)
  7) 체크포인트 — 기업 단위 즉시 checkpoints/results.jsonl append
  8) 반영     — dry-run 아니면 excel_updater.apply_results
"""
from __future__ import annotations

import csv
import datetime as dt
import json
import logging
import os
from dataclasses import dataclass, asdict

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import config
from ai.gemini import GeminiClient, BudgetExceeded, AllModelsExhausted


class VerifyModelsExhausted(Exception):
    """제한된 2단계 검증 모델이 모든 API 키에서 소진됨."""
from collectors import (naver_search, news_search, site_search,
                        thevc_collector, url_collector)
from extractors import investment_extractor as ext
from validators import stage_validator as sv
from updater import excel_updater

log = logging.getLogger(__name__)


@dataclass
class Company:
    row: int
    bizno: str
    name_kr: str
    name_en: str
    industry: str
    stage: str
    htype: str
    website: str

    @property
    def priority(self) -> int:
        return config.priority_of(self.htype)


# ---------------------------------------------------------------- 로드/선정
def load_companies(excel_path=None) -> list[Company]:
    excel_path = excel_path or config.EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[config.SHEET_ALL]

    idx = {k: column_index_from_string(v) - 1 for k, v in {
        "bizno": config.COL_BIZNO, "name_kr": config.COL_NAME_KR,
        "name_en": config.COL_NAME_EN, "industry": config.COL_INDUSTRY,
        "stage": config.COL_STAGE, "htype": config.COL_HTYPE,
        "website": config.COL_WEBSITE,
    }.items()}

    companies = []
    for i, row in enumerate(
        ws.iter_rows(min_row=config.DATA_START_ROW, values_only=True),
        start=config.DATA_START_ROW,
    ):
        def cell(key):
            v = row[idx[key]] if idx[key] < len(row) else None
            return str(v).strip() if v is not None else ""

        name = cell("name_kr")
        if not name:
            continue
        companies.append(Company(
            row=i, bizno=cell("bizno"), name_kr=name, name_en=cell("name_en"),
            industry=cell("industry"), stage=cell("stage"),
            htype=cell("htype"), website=cell("website"),
        ))
    wb.close()
    log.info("전체 기업 로드: %d개", len(companies))
    return companies


def select_targets(companies: list[Company], priority: int | None = None,
                   excel_path=None) -> list[Company]:
    """제외 규칙 적용 + 우선순위 정렬."""
    logged = excel_updater.read_logged_rows(excel_path)
    targets = [
        c for c in companies
        if c.row not in logged and not sv.is_terminal(c.stage)
        and c.priority in config.TARGET_PRIORITIES
    ]
    log.info("제외 후 대상: %d개 (로그 기록 %d행, 종결 상태 제외, 조사 범위 그룹 %s)",
             len(targets), len(logged), sorted(config.TARGET_PRIORITIES))
    targets.sort(key=lambda c: (c.priority, c.row))
    if priority is not None:
        targets = [c for c in targets if c.priority == priority]
        log.info("우선순위 %d 그룹: %d개", priority, len(targets))
    return targets


def load_prescreened() -> dict[int, dict]:
    """스크리닝 기완료 CSV → {row: {verdict, new_stage, evidence, source_url}}"""
    result: dict[int, dict] = {}
    if not config.SCREENED_CSV.exists():
        return result
    with open(config.SCREENED_CSV, newline="", encoding="utf-8-sig") as f:
        for rec in csv.DictReader(f):
            try:
                result[int(rec["row"])] = {
                    "verdict": (rec.get("verdict") or "").strip(),
                    "new_stage": (rec.get("new_stage") or "").strip(),
                    "evidence": (rec.get("evidence") or "").strip(),
                    "source_url": (rec.get("source_url") or "").strip(),
                }
            except (KeyError, TypeError, ValueError):
                continue
    log.info("스크리닝 기완료: %d개", len(result))
    return result


# ---------------------------------------------------------------- 체크포인트
def load_checkpoint() -> dict[int, dict]:
    done: dict[int, dict] = {}
    corrupt = []
    valid_lines = []
    if config.CHECKPOINT_PATH.exists():
        with open(config.CHECKPOINT_PATH, encoding="utf-8") as f:
            for line_no, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                    done[int(rec["row"])] = rec
                    valid_lines.append(line)
                except (json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
                    corrupt.append({"line": line_no, "error": str(e), "raw": line})
                    log.warning("체크포인트 %d행 손상 — 해당 행을 건너뜁니다", line_no)
    if corrupt:
        corrupt_path = config.CHECKPOINT_PATH.with_name("results.corrupt.jsonl")
        with open(corrupt_path, "w", encoding="utf-8") as f:
            for rec in corrupt:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
        # 손상 줄을 제거한 임시 파일로 원자 교체해 이후 append가 잘린 줄에 붙지 않게 한다.
        repair_path = config.CHECKPOINT_PATH.with_suffix(".jsonl.repair")
        with open(repair_path, "w", encoding="utf-8") as f:
            for line in valid_lines:
                f.write(line + "\n")
            f.flush()
            os.fsync(f.fileno())
        os.replace(repair_path, config.CHECKPOINT_PATH)
        log.warning("손상 체크포인트 %d행 별도 보관: %s", len(corrupt), corrupt_path)
    return done


def append_checkpoint(rec: dict):
    config.CHECKPOINT_DIR.mkdir(exist_ok=True)
    with open(config.CHECKPOINT_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")
        f.flush()
        os.fsync(f.fileno())


def export_review_reports() -> tuple:
    """체크포인트에서 반영/확인필요 검수 CSV를 UTF-8-BOM으로 생성한다."""
    import csv
    def actually_applied(rec):
        return rec.get("excel_applied") is True if "excel_applied" in rec \
            else bool(rec.get("applied"))

    headers = ["행", "회사명", "기존 스테이지", "조사 스테이지", "신뢰도",
               "근거", "인용문", "출처 URL", "비고"]
    mappings = [
        (config.CHECKPOINT_DIR / "review_반영.csv", actually_applied),
        (config.CHECKPOINT_DIR / "review_확인필요.csv",
         lambda r: not actually_applied(r)
         and r.get("verdict") in ("changed", "ambiguous")),
    ]
    records = list(load_checkpoint().values())
    config.CHECKPOINT_DIR.mkdir(exist_ok=True)
    counts = []
    for path, predicate in mappings:
        selected = [r for r in records if predicate(r)]
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for r in selected:
                note = "; ".join(x for x in (r.get("note", ""), r.get("apply_error", "")) if x)
                writer.writerow([r.get("excel_row", r.get("row", "")), r.get("name_kr", ""),
                                 r.get("old_stage", ""), r.get("new_stage", ""),
                                 r.get("confidence", ""), r.get("evidence", ""),
                                 r.get("round_quote", ""), r.get("source_url", ""),
                                 note])
        counts.append(len(selected))
    return mappings[0][0], counts[0], mappings[1][0], counts[1]


# ---------------------------------------------------------------- 조사 단계
def _screen_search(c: Company):
    """1단계 검색만 수행 (무료). (query, 병합·정렬된 결과) 반환."""
    query = url_collector.screening_query(c)
    results = news_search.merge_results(
        news_search.search_news(query, max_items=8),
        naver_search.search_news(query, max_items=8),
    )
    return query, results


def _screen_from_results(client, c, query, results) -> ext.ScreeningResult:
    """이미 확보한 검색 결과로 개별 스크리닝 (배치 누락분 폴백에도 사용)."""
    prompt = ext.build_screening_prompt_rss(c, query, news_search.format_block(results))
    answer = client.plain(prompt, model=config.MODEL_SCREEN)
    try:
        return ext.parse_screening(answer)
    except (ValueError, json.JSONDecodeError) as e:
        log.warning("[row %d %s] 스크리닝 파싱 실패 → ambiguous: %s", c.row, c.name_kr, e)
        return ext.ScreeningResult(verdict="ambiguous", note=f"파싱 실패: {e}")


def screen_company(client: GeminiClient, c: Company,
                   search_cache: dict | None = None) -> ext.ScreeningResult:
    if config.SEARCH_MODE == "rss":
        query, results = _screen_search(c)
        if search_cache is not None:
            search_cache[c.row] = results
        if not results:
            # 기사 자체가 없으면 Gemini 호출 없이 종료 (무료 한도 절약)
            return ext.ScreeningResult(verdict="no_info", note="뉴스 검색 결과 없음")
        return _screen_from_results(client, c, query, results)

    prompt = ext.build_screening_prompt(c, url_collector.screening_query(c))
    answer = client.grounded(prompt, model=config.MODEL_SCREEN)
    try:
        return ext.parse_screening(answer)
    except (ValueError, json.JSONDecodeError) as e:
        log.warning("[row %d %s] 스크리닝 파싱 실패 → ambiguous: %s", c.row, c.name_kr, e)
        return ext.ScreeningResult(verdict="ambiguous", note=f"파싱 실패: {e}")


def _screen_pre(c: Company, prescreened: dict):
    """사전 스크리닝(CSV) 결과가 있으면 (ScreeningResult, source_url), 없으면 None."""
    pre = prescreened.get(c.row)
    if pre and pre["verdict"]:
        scr = ext.ScreeningResult(
            verdict=pre["verdict"], new_stage=pre["new_stage"],
            note=f"사전 스크리닝: {pre['evidence']}".strip(": "))
        return scr, pre["source_url"]
    return None


def screen_chunk(client: GeminiClient, chunk: list, prescreened: dict,
                 search_cache: dict | None = None) -> dict:
    """여러 회사를 한 번의 요청으로 스크리닝 → {row: (ScreeningResult, pre_url)}.

    - 사전 스크리닝분과 검색결과 0건은 API 호출 없이 처리
    - 나머지는 배치 1회 호출, 응답 누락분만 개별 폴백
    """
    out: dict = {}
    to_query = []  # (company, query, results)
    for c in chunk:
        pre = _screen_pre(c, prescreened)
        if pre:
            out[c.row] = pre
            continue
        query, results = _screen_search(c)
        if search_cache is not None:
            search_cache[c.row] = results
        if not results:
            out[c.row] = (ext.ScreeningResult(verdict="no_info", note="뉴스 검색 결과 없음"), "")
            continue
        to_query.append((c, query, results))

    if to_query:
        items = [{"id": i, "company": c, "query": q,
                  "search_block": news_search.format_block(r[:5], with_desc=False)}
                 for i, (c, q, r) in enumerate(to_query, 1)]
        prompt = ext.build_screening_prompt_batch(items)
        answer = client.plain(prompt, model=config.MODEL_SCREEN)  # 실패 시 상위로 전파
        id_map = ext.parse_screening_batch(answer, len(items))
        for i, (c, q, r) in enumerate(to_query, 1):
            scr = id_map.get(i)
            if scr is None:
                # 배치 응답 누락 → 캐시된 검색결과로 개별 재조사 (정확성 우선)
                log.info("[row %d %s] 배치 응답 누락 — 개별 재조사", c.row, c.name_kr)
                scr = _screen_from_results(client, c, q, r)
            out[c.row] = (scr, "")
    return out


def verify_company(client: GeminiClient, c: Company,
                   hint: str, cached_results: list | None = None) -> ext.VerificationResult:
    allowed = config.VERIFY_MODELS or None
    if not client.has_available_models(allowed, config.MODEL_VERIFY):
        raise VerifyModelsExhausted("제한된 검증 모델이 모든 API 키에서 소진됨")
    if config.SEARCH_MODE == "rss":
        bare = bool(config.BARE_TERMINAL_RE.match(c.stage))
        pools = [cached_results or []]
        for q in url_collector.verification_queries_rss(c, bare_terminal=bare):
            pools.append(news_search.search_news(q, max_items=6))
            pools.append(naver_search.search_news(q, max_items=6))
        merged = news_search.merge_results(*pools, cap=12)  # 중복 제거 + 최신 우선
        articles = site_search.collect_articles(c.name_kr)  # 기사 본문 확보
        thevc = thevc_collector.get_block(c.name_kr) if config.THEVC_ENABLED else ""
        prompt = ext.build_verification_prompt_rss(
            c, hint, news_search.format_block(merged[:12]), thevc,
            site_search.format_block(articles))
        try:
            answer = client.plain(prompt, model=config.MODEL_VERIFY,
                                  allowed_models=allowed)
        except AllModelsExhausted as e:
            raise VerifyModelsExhausted(str(e)) from e
    else:
        queries = url_collector.verification_queries(c, hint_stage=hint)
        prompt = ext.build_verification_prompt(c, hint, " / ".join(queries))
        try:
            answer = client.grounded(prompt, model=config.MODEL_VERIFY,
                                     allowed_models=allowed)
        except AllModelsExhausted as e:
            raise VerifyModelsExhausted(str(e)) from e

    try:
        return ext.parse_verification(answer)
    except (ValueError, json.JSONDecodeError) as e:
        log.warning("[row %d %s] 검증 파싱 실패: %s", c.row, c.name_kr, e)
        return ext.VerificationResult(
            new_stage="", confidence="low", evidence="", note=f"파싱 실패: {e}",
        )


# ---------------------------------------------------------------- 실행
def run(priority: int | None = None, limit: int | None = None,
        dry_run: bool = False, max_calls: int | None = None,
        batch_size: int | None = None, excel_path=None) -> list[dict]:
    excel_path = excel_path or config.EXCEL_PATH
    companies = load_companies(excel_path)
    targets = select_targets(companies, priority, excel_path)
    prescreened = load_prescreened()
    done = load_checkpoint()

    if done:
        before = len(targets)
        targets = [c for c in targets if c.row not in done]
        log.info("체크포인트 스킵: %d개 (남은 대상 %d개)", before - len(targets), len(targets))

    if limit:
        targets = targets[:limit]

    if batch_size is None:
        batch_size = config.SCREEN_BATCH_SIZE
    client = GeminiClient(max_calls=max_calls)
    results: list[dict] = []
    search_cache: dict[int, list] = {}
    stats = {"changed": 0, "unchanged": 0, "no_info": 0, "ambiguous": 0,
             "applied": 0, "verified": 0}

    use_batch = config.SEARCH_MODE == "rss" and batch_size > 1
    log.info("조사 시작: %d개 (dry-run=%s, max-calls=%s, 배치=%s)",
             len(targets), dry_run, max_calls, batch_size if use_batch else "off")

    if use_batch:
        _run_batched(client, targets, prescreened, results, stats, batch_size, search_cache)
    else:
        _run_serial(client, targets, prescreened, results, stats, search_cache)

    log.info("=== 요약: 조사 %d / changed %d / unchanged %d / no_info %d / "
             "ambiguous %d / 2단계 검증 %d / 반영 %d / API 호출 %d회 ===",
             len(results), stats["changed"], stats["unchanged"], stats["no_info"],
             stats["ambiguous"], stats["verified"], stats["applied"],
             client.call_count)

    if not dry_run:
        # 이번 세션 결과만이 아니라 체크포인트 전체를 반영 — 언제 실행해도
        # 출력 파일은 항상 '원본 + 지금까지의 모든 조사 결과'가 되도록 (누적/멱등)
        all_records = list(load_checkpoint().values())
        if all_records:
            group_desc = f"우선순위 {priority} 그룹" if priority else "전체 우선순위"
            before_apply = {r["row"]: (r.get("excel_applied"), r.get("apply_error"),
                                        r.get("excel_row"))
                            for r in all_records}
            excel_updater.apply_results(
                all_records, f"{group_desc} (누적 {len(all_records)}건)", excel_path)
            for rec in all_records:
                outcome = (rec.get("excel_applied"), rec.get("apply_error"),
                           rec.get("excel_row"))
                if outcome != before_apply.get(rec["row"]):
                    append_checkpoint(rec)
    else:
        log.info("dry-run — 엑셀 반영 생략 (체크포인트에만 기록)")
    return results


def _record(rec: dict, c: Company, results: list):
    """레코드를 결과 목록에 추가 + 체크포인트 즉시 저장 + 한 줄 로그."""
    results.append(rec)
    append_checkpoint(rec)
    log.info("[row %d] %s | %s | %s → %s (%s)",
             c.row, c.name_kr, rec["verdict"], rec["old_stage"] or "(공란)",
             rec["new_stage"] or "-", rec["status"])


def _run_serial(client, targets, prescreened, results, stats, search_cache=None):
    """회사별 순차 처리 (grounding 모드 또는 배치 미사용 시)."""
    for c in targets:
        try:
            rec = _process_one(client, c, prescreened, stats, search_cache)
        except BudgetExceeded as e:
            log.warning("비용 가드 발동: %s — 저장 후 정상 종료", e)
            return
        except VerifyModelsExhausted as e:
            log.warning("[row %d %s] 검증 모델 소진 — 체크포인트 없이 다음 실행에서 재시도: %s",
                        c.row, c.name_kr, e)
            continue
        except Exception as e:
            log.error("[row %d %s] API 오류로 중단 — 재실행하면 이어서 진행됩니다: %s",
                      c.row, c.name_kr, e)
            return
        _record(rec, c, results)


def _run_batched(client, targets, prescreened, results, stats, batch_size,
                 search_cache=None):
    """1단계 스크리닝을 batch_size개씩 묶어 처리, 2단계 검증은 회사별 개별."""
    for i in range(0, len(targets), batch_size):
        chunk = targets[i:i + batch_size]
        try:
            screened = screen_chunk(client, chunk, prescreened, search_cache)
        except BudgetExceeded as e:
            log.warning("비용 가드 발동: %s — 저장 후 정상 종료", e)
            return
        except Exception as e:
            log.error("스크리닝 배치 중단 — 재실행하면 이어서 진행됩니다: %s", e)
            return
        for c in chunk:
            scr, pre_url = screened[c.row]
            try:
                rec = _finalize(client, c, scr, pre_url, stats,
                                (search_cache or {}).get(c.row))
            except BudgetExceeded as e:
                log.warning("비용 가드 발동: %s — 저장 후 정상 종료", e)
                return
            except VerifyModelsExhausted as e:
                log.warning("[row %d %s] 검증 모델 소진 — 체크포인트 없이 다음 실행에서 재시도: %s",
                            c.row, c.name_kr, e)
                continue
            except Exception as e:
                log.error("[row %d %s] 검증 중단 — 재실행하면 이어서 진행됩니다: %s",
                          c.row, c.name_kr, e)
                return
            _record(rec, c, results)


def _process_one(client: GeminiClient, c: Company,
                 prescreened: dict[int, dict], stats: dict,
                 search_cache: dict | None = None) -> dict:
    # ---- 1단계 스크리닝 (기완료분은 스킵)
    pre = _screen_pre(c, prescreened)
    if pre:
        scr, pre_url = pre
    else:
        scr = screen_company(client, c, search_cache)
        pre_url = ""
    return _finalize(client, c, scr, pre_url, stats,
                     (search_cache or {}).get(c.row))


def _finalize(client: GeminiClient, c: Company, scr: ext.ScreeningResult,
              pre_url: str, stats: dict, cached_results: list | None = None) -> dict:
    """스크리닝 결과 → (필요 시) 2단계 검증 → 반영 판정 → 체크포인트 레코드."""
    stats[scr.verdict] = stats.get(scr.verdict, 0) + 1
    rec = {
        "ts": dt.datetime.now().isoformat(timespec="seconds"),
        "row": c.row, "bizno": c.bizno, "name_kr": c.name_kr,
        "priority": c.priority, "old_stage": c.stage,
        "verdict": scr.verdict, "new_stage": "", "confidence": "",
        "evidence": "", "source_url": "", "note": scr.note,
        "applied": False, "apply_decision": False, "excel_applied": False,
        "apply_error": "", "excel_row": "", "status": "",
    }

    # 연도 없는 M&A/IPO 는 verdict와 무관하게 연도 보정 필요 → 검증 대상
    needs_verify = scr.verdict in ("changed", "ambiguous") or \
        bool(config.BARE_TERMINAL_RE.match(c.stage))

    if not needs_verify:
        rec["status"] = "변경 없음" if scr.verdict == "unchanged" else "미반영(정보 없음)"
        return rec

    # ---- 2단계 정밀 검증
    hint = scr.new_stage or scr.note
    ver = verify_company(client, c, hint, cached_results)
    stats["verified"] += 1

    # 라운드명 인용 검사 — 인용문에 라운드명이 없으면 기계적으로 신뢰도 강등
    confidence, demote_reason = sv.check_round_quote(
        ver.new_stage, ver.round_quote, ver.confidence)

    decision = sv.decide(c.stage, ver.new_stage, confidence, ver.note)
    rec.update({
        "new_stage": sv.normalize_stage(ver.new_stage) or ver.new_stage,
        "confidence": confidence,
        "evidence": ver.evidence,
        "round_quote": ver.round_quote,
        "article_date": ver.article_date,
        "source_url": ver.source_url or pre_url,
        "note": "; ".join(x for x in (rec["note"], ver.note, demote_reason) if x),
        "applied": decision.apply,
        "apply_decision": decision.apply,
        "status": decision.status,
    })
    if decision.apply:
        stats["applied"] += 1
    return rec
