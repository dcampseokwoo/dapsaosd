"""US FORGED — DB 스냅샷 로더 + provenance (§7 기초).

살아있는 구글 시트에서 export되는 DB는 실행마다 바뀐다(18시간 만에 스테이지 652곳 변동).
그래서 입력을 data/snapshots/ 에 고정한 xlsx 로만 읽고, 실행에 쓴 스냅샷의
파일명·SHA256·행수·스테이지 분포를 provenance 로 남긴다(리포트 기록용, assertion 아님).

이 모듈은 로드·정규화·인덱싱만 한다. 분류/스테이지/배제 판정은 각 레이어 모듈이 한다.
PII(대표 이메일·연락처)는 내부 행에는 담되(§6 연락처 분리에 필요) 산출물에서 제외한다.
"""
from __future__ import annotations

import hashlib
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SNAP_DIR = ROOT / "data" / "snapshots"
DEFAULT_SNAPSHOT = SNAP_DIR / "GBD_DB_20260818_1856.xlsx"

# DB 헤더(3행) → 내부 키
_COLMAP = {
    "키값(사업자등록번호)": "biz_no", "국문 회사명": "name_ko", "영문 회사명": "name_en",
    "업종(CB 기준)": "industry", "기술": "tech", "투자 스테이지": "stage",
    "재단연관기업 분류": "foundation_type", "투자사 또는 펀드명": "investor",
    "타겟 국가": "target", "영문 서비스명": "svc", "1줄 사업 소개": "desc",
    "Website": "website", "대표자 성함": "ceo", "대표자 이메일": "email",
    "대표자 연락처": "phone", "비고": "note", "GBD 프로그램 참여 이력('25~)": "program_history",
}

_BIZ_VALID = re.compile(r"^\d{3}-\d{2}-\d{5}$")
_BIZ_FOREIGN = re.compile(r"^(OC\w+|외국법인[_\s].*|해외법인.*)$", re.IGNORECASE)


def normalize_biz_no(value) -> tuple[str, str]:
    """사업자번호 정규화 → (정규화값, 상태).

    상태: valid(\\d3-\\d2-\\d5) / foreign(OC*·외국법인_*·해외법인) / malformed / empty.
    공백 제거 후 판정. 하이픈 위치 오류('725-870-2428')는 malformed 로 표시(조용히 버리지 않음).
    """
    if value is None:
        return "", "empty"
    s = str(value).strip()
    if not s:
        return "", "empty"
    if _BIZ_VALID.match(s):
        return s, "valid"
    if _BIZ_FOREIGN.match(s):
        return s, "foreign"
    # 숫자·하이픈만인데 패턴 불일치 → 자릿수/하이픈 오류
    digits = re.sub(r"\D", "", s)
    if len(digits) == 10:
        canon = f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
        return canon, "malformed"     # 정규화는 하되 malformed 로 리포트
    return s, "malformed"


def load_rows(path: Path | str = DEFAULT_SNAPSHOT) -> list[dict]:
    """스냅샷 xlsx → 정규화된 행 리스트. 국문 회사명이 있는 행만."""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["All(전체기업)"] if "All(전체기업)" in wb.sheetnames else wb.worksheets[1]
    header_row = 3
    hdr = {}
    for c, cell in enumerate(next(ws.iter_rows(min_row=header_row, max_row=header_row,
                                               values_only=True)), 1):
        if cell in _COLMAP:
            hdr[_COLMAP[cell]] = c
    rows = []
    for raw in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def g(key):
            i = hdr.get(key)
            v = raw[i - 1] if i and i - 1 < len(raw) else None
            return "" if v is None else str(v).strip()
        if not g("name_ko"):
            continue
        biz, biz_status = normalize_biz_no(g("biz_no"))
        rows.append({
            "biz_no": biz, "biz_status": biz_status, "biz_no_raw": g("biz_no"),
            "name_ko": g("name_ko"), "name_en": g("name_en"), "svc": g("svc"),
            "industry": g("industry"), "tech": g("tech"), "stage": g("stage"),
            "foundation_type": g("foundation_type"), "investor": g("investor"),
            "target": g("target"), "desc": g("desc"), "website": g("website"),
            "email": g("email"), "phone": g("phone"), "note": g("note"),
            "program_history": g("program_history"),
        })
    wb.close()
    return rows


def index_by_biz(rows: list[dict]) -> dict[str, list[dict]]:
    """정규화 사업자번호 → 행들(중복 가능). 빈 biz_no 는 제외."""
    idx: dict[str, list[dict]] = {}
    for r in rows:
        if r["biz_no"]:
            idx.setdefault(r["biz_no"], []).append(r)
    return idx


def sha256(path: Path | str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def provenance(path: Path | str = DEFAULT_SNAPSHOT, rows: list[dict] | None = None) -> dict:
    """실행에 쓴 스냅샷 메타(리포트 기록용). run_timestamp 는 호출측에서 주입."""
    from collections import Counter
    path = Path(path)
    rows = rows if rows is not None else load_rows(path)
    return {
        "input_snapshot": path.name,
        "input_sha256": sha256(path),
        "input_rows": len(rows),
        "stage_dist": dict(Counter(r["stage"] or "(빈값)" for r in rows).most_common()),
        "biz_status_dist": dict(Counter(r["biz_status"] for r in rows).most_common()),
    }
