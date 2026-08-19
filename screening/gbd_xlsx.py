"""GBD 대규모 자동 평가 → 엑셀. output/screening/gbd_auto_eval.xlsx

시트: 1) 요약·퍼널  2) 재단분류 교차표  3) 전체 판정(3,424개사)  4) 방법론
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from screening import rules

OUT = Path(__file__).resolve().parent.parent / "output" / "screening" / "gbd_auto_eval.xlsx"
FONT = "Arial"
HDR = PatternFill("solid", fgColor="1F3864")
HF = Font(name=FONT, bold=True, color="FFFFFF", size=10)
SUB = PatternFill("solid", fgColor="D6DCE4")
GREEN = PatternFill("solid", fgColor="E2EFDA")
YEL = PatternFill("solid", fgColor="FFF2CC")
ORG = PatternFill("solid", fgColor="FCE4D6")
GRY = PatternFill("solid", fgColor="EDEDED")
THIN = Side(style="thin", color="BFBFBF")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _c(ws, r, c, v, *, bold=False, fill=None, wrap=False, align="left", size=10):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=FONT, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORD
    if fill:
        cell.fill = fill
    return cell


def _hdr(ws, r, hs, ws_):
    for i, (h, w) in enumerate(zip(hs, ws_), 1):
        c = _c(ws, r, i, h, bold=True, fill=HDR, align="center")
        c.font = HF
        ws.column_dimensions[get_column_letter(i)].width = w


def _fill_for(outcome: str) -> PatternFill:
    if "탈락" in outcome or "라우팅" in outcome:
        return ORG
    if "사람 검토" in outcome:
        return YEL
    if "통과" in outcome:
        return GREEN
    return GRY   # 미상/판정불가


def build(rows, s) -> Path:
    wb = Workbook()

    # ---- 시트1 요약 ----
    ws = wb.active
    ws.title = "요약_퍼널"
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, f"디캠프 GBD 스타트업 마스터 DB — 500/HAX 엔진 대규모 자동 평가 "
       f"({s['n']:,}개사)", bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _c(ws, 2, 1, "DB 정형 필드(업종·기술·스테이지·1줄소개·재단분류)만으로 1차 필터"
       "(트랙 라우팅+하드 게이트)를 결정적 규칙으로 자동 적용. 개별 크롤링·LLM 호출 "
       "없음. 4축 점수(v2/v3)는 1줄 소개만으로 신뢰성 있게 못 매기므로 이 대규모 "
       "단계에서는 산출하지 않는다(팩트시트 있는 102개사는 engine_full_eval.xlsx).",
       size=9, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 46

    r = 4
    _c(ws, r, 1, "트랙 라우팅", bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    _hdr(ws, r, ["트랙", "기업 수"], [34, 12])
    r += 1
    for k, n in s["track"].most_common():
        _c(ws, r, 1, k)
        _c(ws, r, 2, n, align="center")
        r += 1
    r += 1
    _c(ws, r, 1, "1차 필터 결과", bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    _hdr(ws, r, ["게이트 판정", "기업 수"], [34, 12])
    r += 1
    for k, n in s["outcome"].most_common():
        _c(ws, r, 1, k, fill=_fill_for(k))
        _c(ws, r, 2, n, align="center", fill=_fill_for(k))
        r += 1
    scoreable = sum(n for k, n in s["outcome"].items() if "통과" in k)
    r += 1
    _c(ws, r, 1, f"→ 점수화 대상(게이트 통과): {scoreable:,}개사 / "
       f"사람 검토 {s['outcome'].get('사람 검토 (500: A 이후 밴드)', 0):,} / "
       f"게이트 탈락 {s['outcome'].get('게이트 탈락 (HAX 스테이지 이탈 — 시리즈A+)', 0):,}",
       bold=True, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28

    # ---- 시트2 재단분류 교차표 ----
    ws2 = wb.create_sheet("재단분류_교차표")
    ws2.sheet_view.showGridLines = False
    _c(ws2, 1, 1, "재단연관 분류 × 1차 필터 결과 (Type1 패밀리사 / Type2·3 포트폴리오 / 디데이)",
       bold=True, size=12)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    cats = ["Type 1", "Type 2", "Type 3", "디데이 출전팀"]
    buckets = ["점수화 통과", "사람 검토", "게이트 탈락", "바이오 라우팅",
               "스테이지 미상", "정보 부족"]

    def bucket(o):
        if "통과" in o:
            return "점수화 통과"
        if "사람 검토" in o:
            return "사람 검토"
        if "게이트 탈락" in o:
            return "게이트 탈락"
        if "IndieBio" in o:
            return "바이오 라우팅"
        if "미상" in o:
            return "스테이지 미상"
        return "정보 부족"

    cross = {c: Counter() for c in cats}
    for row in rows:
        t = row["type"].split(",")[0].split("(")[0].strip()
        if t in cross:
            cross[t][bucket(row["outcome"])] += 1
    _hdr(ws2, 3, ["재단분류"] + buckets + ["합계"], [16] + [12] * 6 + [10])
    rr = 4
    for c in cats:
        _c(ws2, rr, 1, c, bold=True)
        tot = 0
        for j, b in enumerate(buckets, 2):
            n = cross[c][b]
            tot += n
            _c(ws2, rr, j, n or "", align="center",
               fill=GREEN if b == "점수화 통과" and n else
               YEL if b == "사람 검토" and n else
               ORG if b in ("게이트 탈락", "바이오 라우팅") and n else None)
        _c(ws2, rr, 8, tot, align="center", bold=True)
        rr += 1
    rr += 1
    _c(ws2, rr, 1, "읽는 법: Type1 패밀리사는 디캠프가 직접 투자·입주시킨 기업이라 "
       "스테이지·업종 정보가 충실해 라우팅·게이트가 잘 정해진다. Type3(하위펀드)은 "
       "정보가 얕아 '정보 부족'이 많다. 이는 DB 완결성의 문제이지 엔진의 문제가 아니다.",
       size=9, wrap=True)
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
    ws2.row_dimensions[rr].height = 40

    # ---- 시트3 전체 판정 ----
    ws3 = wb.create_sheet("전체_판정")
    ws3.sheet_view.showGridLines = False
    _c(ws3, 1, 1, f"전체 판정표 ({s['n']:,}개사) — DB 필드 기반 자동 라우팅·게이트",
       bold=True, size=12)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _hdr(ws3, 3, ["국문명", "영문명", "업종(CB)", "스테이지", "재단분류",
                  "→ 트랙", "밴드", "1차 필터 결과"],
         [20, 18, 20, 12, 18, 12, 10, 32])
    rr = 4
    order = {"통과": 0, "사람 검토": 1, "탈락": 2, "IndieBio": 3, "미상": 4, "부족": 5}

    def okey(row):
        o = row["outcome"]
        for k, v in order.items():
            if k in o:
                return v
        return 9
    for row in sorted(rows, key=lambda r: (okey(r), r["track"])):
        f = _fill_for(row["outcome"])
        _c(ws3, rr, 1, row["name_ko"], fill=f, size=9)
        _c(ws3, rr, 2, row["name_en"], fill=f, size=8)
        _c(ws3, rr, 3, row["sector"][:40], fill=f, size=8)
        _c(ws3, rr, 4, row["stage"], fill=f, size=8, align="center")
        _c(ws3, rr, 5, row["type"][:26], fill=f, size=8)
        _c(ws3, rr, 6, row["track"], fill=f, size=9, align="center")
        _c(ws3, rr, 7, row["band"], fill=f, size=8, align="center")
        _c(ws3, rr, 8, row["outcome"], fill=f, size=8)
        rr += 1
    ws3.freeze_panes = "A4"

    # ---- 시트4 방법론 ----
    ws4 = wb.create_sheet("방법론_한계")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 115
    notes = [
        ("방법론 및 한계 — 대규모 자동 평가", True, 12),
        ("데이터 출처", True, 11),
        ("디캠프 GBD 스타트업 데이터베이스 Ver.26.01 (전체 3,424개사). 대표 이메일·"
         "연락처 등 개인정보는 추출 단계에서 제외했다.", False, 10),
        ("자동화한 것 / 못 한 것", True, 11),
        ("자동화 O: 트랙 라우팅(업종·기술·1줄소개 키워드 → 500/HAX/IndieBio)과 하드 "
         "게이트(투자 스테이지 → 통과/사람검토/탈락)를 결정적 규칙으로 전 기업에 적용. "
         "개별 웹 크롤링·LLM 호출 없이 즉시 실행된다.", False, 10),
        ("자동화 X: 4축 레벨 점수(v2 Tier·v3 구간)는 1줄 소개만으로 Traction/Team/"
         "Market/Moat 를 신뢰성 있게 매길 수 없어 이 단계에서는 내지 않는다. 점수는 "
         "팩트시트를 수집한 102개사(engine_full_eval.xlsx)에서만 산출했다.", False, 10),
        ("라우팅 규칙", True, 11),
        ("① 신약·치료제·항체·백신 등 바이오 치료제 키워드 → IndieBio 라우팅(진단·"
         "의료기기·디지털헬스는 제외). ② 로봇·소재·배터리·센서·제조·에너지·우주 등 "
         "하드웨어 키워드 → HAX. ③ 그 외 → 500(섹터 무관). 업종·기술·소개가 모두 "
         "공란이면 '판정 불가(정보 부족)'.", False, 10),
        ("게이트 규칙", True, 11),
        ("시리즈A 이상 → HAX 는 탈락(프리시드~시드 대상), 500 은 사람 검토. "
         "Seed/Pre-A/Angel → 점수화 대상. 스테이지 미상 → 판정 보류(자료 요청).", False, 10),
        ("이 결과로 말할 수 있는 것 / 없는 것", True, 11),
        ("● 말할 수 있는 것: 엔진의 1차 필터가 3,424개 모집단에서 어떻게 분포하는지 — "
         "트랙 배분, 스테이지 게이트 통과·탈락 비율, 재단분류별 차이. 정형 데이터만으로 "
         "라우팅·게이트가 대규모 자동화된다.", False, 10),
        ("● 말할 수 없는 것: 개별 기업의 합불·점수. 키워드 라우팅은 근사이며 경계 "
         "사례(HW+SW 혼합, 진단 vs 치료제)는 오분류가 있을 수 있다. '정보 부족'·"
         "'스테이지 미상'이 많은 것은 DB 완결성의 문제다.", False, 10),
    ]
    for i, (t, b, sz) in enumerate(notes, 1):
        c = _c(ws4, i, 1, t, bold=b, size=sz, wrap=True)
        if b and sz >= 11:
            c.fill = SUB
        ws4.row_dimensions[i].height = 44 if len(t) > 70 else 16

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


OUT_V4 = OUT.parent / "gbd_auto_eval_v4.xlsx"

V4_ORDER = ["점수화 대상", "사람 검토 (경계 스테이지/보류)",
            "스케일업 트랙 안내 (스테이지 명백 이탈)", "라우팅 사람 확인 (신호 접전)",
            "IndieBio 라우팅", "자료 요청 (스테이지 미상)", "자료 요청 (트랙 특정 불가)",
            "평가 대상외 (입력 없음)"]
V4_FILL = {"점수화 대상": GREEN, "사람 검토 (경계 스테이지/보류)": YEL,
           "스케일업 트랙 안내 (스테이지 명백 이탈)": PatternFill("solid", fgColor="DEEBF7"),
           "라우팅 사람 확인 (신호 접전)": ORG, "IndieBio 라우팅": ORG,
           "자료 요청 (스테이지 미상)": GRY, "자료 요청 (트랙 특정 불가)": GRY,
           "평가 대상외 (입력 없음)": GRY}


def build_v4(v4_rows, v3_rows) -> Path:
    """v4 재설계 워크북 — v3 대비 비교 포함."""
    from collections import Counter
    wb = Workbook()

    # 시트1 요약 + v3 대비
    ws = wb.active
    ws.title = "요약_v4"
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, f"GBD 마스터 DB — 엔진 v4 대규모 자동 평가 ({len(v4_rows):,}개사)",
       bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    _c(ws, 2, 1, "v4 재설계: 다신호 가중 라우터(자기불확실성 플래그) + 스테이지 3분할 "
       "게이트(시드=점수화 / 시리즈A=경계 사람검토 / 시리즈B+=스케일업 안내) + 입력 "
       "상태 분리. WEIGHTS·컷오프 불변. 4축 점수는 팩트시트 있는 102개사에만.",
       size=9, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.row_dimensions[2].height = 40

    def v3bucket(o):
        if "통과" in o:
            return "점수화 대상"
        if "사람 검토" in o:
            return "사람 검토 (경계 스테이지/보류)"
        if "게이트 탈락" in o:
            return "게이트 탈락(v3) → 스케일업 재분류"
        if "IndieBio" in o:
            return "IndieBio 라우팅"
        if "미상" in o:
            return "자료 요청 (스테이지 미상)"
        return "평가 대상외 (입력 없음)"
    c3 = Counter(v3bucket(r["outcome"]) for r in v3_rows)
    c4 = Counter(r["outcome"] for r in v4_rows)

    _c(ws, 4, 1, "1차 필터 결과 — v3(기존) → v4(재설계)", bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
    _hdr(ws, 5, ["v4 버킷", "v4 기업 수", "(참고) v3 동류", "v3 수"], [36, 12, 30, 10])
    rr = 6
    for b in V4_ORDER:
        _c(ws, rr, 1, b, fill=V4_FILL.get(b))
        _c(ws, rr, 2, c4.get(b, 0), align="center", bold=True, fill=V4_FILL.get(b))
        rr += 1
    _c(ws, rr, 1, "── v3 참고: 사람 검토 468 / 게이트 탈락 225 → v4에서 스케일업 안내로 "
       "재분류, 사람 검토 반감 ──", size=9, wrap=True)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    ws.row_dimensions[rr].height = 26
    rr += 2
    delta = [("사람 검토", c3.get("사람 검토 (경계 스테이지/보류)", 0),
              c4.get("사람 검토 (경계 스테이지/보류)", 0)),
             ("게이트 탈락 → 스케일업 안내", c3.get("게이트 탈락(v3) → 스케일업 재분류", 0),
              c4.get("스케일업 트랙 안내 (스테이지 명백 이탈)", 0)),
             ("라우팅 사람 확인(신규)", 0, c4.get("라우팅 사람 확인 (신호 접전)", 0))]
    _c(ws, rr, 1, "핵심 변화", bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    rr += 1
    _hdr(ws, rr, ["지표", "v3", "v4", "변화"], [30, 10, 10, 24])
    rr += 1
    for lab, a, b in delta:
        _c(ws, rr, 1, lab)
        _c(ws, rr, 2, a, align="center")
        _c(ws, rr, 3, b, align="center", bold=True)
        _c(ws, rr, 4, ("반감" if b < a else "신규 플래그" if a == 0 else "재분류"),
           align="center", size=9)
        rr += 1

    # 시트2 라우팅 불안정(저신뢰) 목록
    ws2 = wb.create_sheet("라우팅_불안정")
    ws2.sheet_view.showGridLines = False
    _c(ws2, 1, 1, "라우팅 불안정 — 신호 접전으로 사람 확인 권장 (조용한 오분류 방지)",
       bold=True, size=12)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    low = [r for r in v4_rows if r.get("route_conf") == "low"
           and r["track"] not in ("대상외", "판정 보류")]
    _c(ws2, 2, 1, f"총 {len(low):,}개사 저신뢰 라우팅. 이 중 hax/bio 로 기운 건 점수화 "
       "전 사람 확인으로 보낸다(500 저신뢰는 섹터 무관 기본값이라 진행).", size=9, wrap=True)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    _hdr(ws2, 4, ["국문명", "업종(CB)", "기술", "→ 트랙", "판정", "라우팅 점수"],
         [20, 20, 16, 10, 26, 30])
    rr = 5
    for r in sorted(low, key=lambda x: x["name_ko"])[:400]:
        _c(ws2, rr, 1, r["name_ko"], size=9)
        _c(ws2, rr, 2, r["sector"][:30], size=8)
        _c(ws2, rr, 3, r["tech"][:22], size=8)
        _c(ws2, rr, 4, r["track"], align="center", size=9)
        _c(ws2, rr, 5, r["outcome"], size=8)
        _c(ws2, rr, 6, r["route_reason"].split("—")[0][-30:], size=8)
        rr += 1

    # 시트3 전체 판정
    ws3 = wb.create_sheet("전체_판정_v4")
    ws3.sheet_view.showGridLines = False
    _c(ws3, 1, 1, f"전체 판정표 v4 ({len(v4_rows):,}개사)", bold=True, size=12)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _hdr(ws3, 3, ["국문명", "영문명", "업종(CB)", "스테이지", "재단분류",
                  "→ 트랙", "밴드", "v4 판정"], [20, 16, 20, 12, 18, 10, 10, 30])
    rr = 4
    ordr = {b: i for i, b in enumerate(V4_ORDER)}
    for r in sorted(v4_rows, key=lambda x: (ordr.get(x["outcome"], 9), x["track"])):
        f = V4_FILL.get(r["outcome"], GRY)
        _c(ws3, rr, 1, r["name_ko"], fill=f, size=9)
        _c(ws3, rr, 2, r["name_en"], fill=f, size=8)
        _c(ws3, rr, 3, r["sector"][:38], fill=f, size=8)
        _c(ws3, rr, 4, r["stage"], fill=f, size=8, align="center")
        _c(ws3, rr, 5, r["type"][:26], fill=f, size=8)
        _c(ws3, rr, 6, r["track"], fill=f, size=9, align="center")
        _c(ws3, rr, 7, r["band"], fill=f, size=8, align="center")
        _c(ws3, rr, 8, r["outcome"], fill=f, size=8)
        rr += 1
    ws3.freeze_panes = "A4"

    # 시트4 방법론
    ws4 = wb.create_sheet("방법론_v4")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 115
    notes = [
        ("엔진 v4 — 무엇이 바뀌었나", True, 12),
        ("1. 다신호 가중 라우터", True, 11),
        ("bio/hw/sw 3축 가중 점수(사업소개 1.0 > 기술태그 0.7 > 섹터 0.5). 1·2위 차가 "
         "작으면 '라우팅 불안정'으로 스스로 플래그 → 조용한 오분류 방지(예: 메텍홀딩스 "
         "tech='소프트웨어'라도 소개가 하드웨어면 접전으로 표시).", False, 10),
        ("2. 스테이지 3분할 게이트", True, 11),
        ("시드=점수화 / 시리즈A=경계 사람검토(아직 지원 여지) / 시리즈B+=스케일업 트랙 "
         "안내(탈락이 아니라 '단계가 다름'). 효과: 사람 검토 468→244 반감, 게이트 탈락 "
         "225→0(스케일업 재분류).", False, 10),
        ("3. 입력 상태 분리", True, 11),
        ("빈 DB 행을 '판정 불가'가 아니라 '평가 대상외(입력 없음)'로 명시. 신호 약함은 "
         "'자료 요청'. 엔진이 판정을 안 하는 게 아니라 판정할 입력이 없는 것.", False, 10),
        ("불변·한계", True, 11),
        ("WEIGHTS·컷오프 불변(라벨 튜닝 없음). 키워드 라우터는 근사이며 v4의 개선은 "
         "'틀릴 때 스스로 플래그'하는 것. 4축 점수는 1줄 소개론 불가 → 팩트시트 있는 "
         "102개사(engine_full_eval.xlsx)에만. 정밀도·특이도는 진짜 불합격 라벨 필요.",
         False, 10),
    ]
    for i, (t, b, sz) in enumerate(notes, 1):
        c = _c(ws4, i, 1, t, bold=b, size=sz, wrap=True)
        if b and sz >= 11:
            c.fill = SUB
        ws4.row_dimensions[i].height = 44 if len(t) > 70 else 16

    OUT_V4.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_V4)
    return OUT_V4


OUT_V5 = OUT.parent / "gbd_auto_eval_v5.xlsx"
V5_ORDER = ["점수화 대상", "조건부 통과 → 점수화 (설문 필요)", "사람 검토 (경계)",
            "확정 탈락", "라우팅 사람 확인 (신호 접전)", "IndieBio 라우팅",
            "자료 요청 (스테이지 미상)", "자료 요청 (트랙 특정 불가)",
            "평가 대상외 (입력 없음)"]
V5_FILL = {"점수화 대상": GREEN, "조건부 통과 → 점수화 (설문 필요)": GREEN,
           "사람 검토 (경계)": YEL, "확정 탈락": PatternFill("solid", fgColor="F4B7B7"),
           "라우팅 사람 확인 (신호 접전)": ORG, "IndieBio 라우팅": ORG,
           "자료 요청 (스테이지 미상)": GRY, "자료 요청 (트랙 특정 불가)": GRY,
           "평가 대상외 (입력 없음)": GRY}


def build_v5(v5_rows, v4_rows) -> Path:
    """v5 확정 탈락 활성 워크북 — v4(물렁) 대비 + 탈락 사유."""
    from collections import Counter
    wb = Workbook()

    # 시트1 요약
    ws = wb.active
    ws.title = "요약_v5"
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, f"GBD 마스터 DB — 엔진 v5 '확정 탈락' 활성 ({len(v5_rows):,}개사)",
       bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    _c(ws, 2, 1, "v5 원칙: '확실히 아닌 건 확실히 탈락.' 프로그램이 절대 waive 안 하는 "
       "기준(스테이지 이탈·섹터 부적합·언어·제품·커밋·오너십)에 **확인된 사실**이 걸리면 "
       "즉시 확정 탈락(사유 명시). 확인 안 된 건 추측 탈락 없이 조건부(설문). "
       "v4의 '스케일업 안내'(물렁 유보)는 철회 → 시리즈B+·HAX 시리즈A 는 확정 탈락.",
       size=9, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.row_dimensions[2].height = 52

    c5 = Counter(r["outcome"] for r in v5_rows)
    _c(ws, 4, 1, "1차 필터 결과 (v5)", bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    _hdr(ws, 5, ["판정", "기업 수"], [40, 12])
    rr = 6
    for b in V5_ORDER:
        _c(ws, rr, 1, b, fill=V5_FILL.get(b))
        _c(ws, rr, 2, c5.get(b, 0), align="center", bold=True, fill=V5_FILL.get(b))
        rr += 1

    # 확정 탈락 사유 분포
    rr += 1
    fails = [r for r in v5_rows if r["outcome"] == "확정 탈락"]
    reasons = Counter()
    for r in fails:
        for x in (r.get("reasons") or "").split("; "):
            if x:
                reasons[x.split(":")[0].split("—")[0].strip()] += 1
    _c(ws, rr, 1, f"확정 탈락 {len(fails):,}개사 — 사유 분포", bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    rr += 1
    _hdr(ws, rr, ["탈락 사유", "기업 수"], [40, 12])
    rr += 1
    for k, n in reasons.most_common():
        _c(ws, rr, 1, k, fill=V5_FILL["확정 탈락"])
        _c(ws, rr, 2, n, align="center")
        rr += 1

    # v4 대비
    rr += 1
    def v4b(o):
        if "스케일업" in o:
            return "스케일업 안내"
        if "사람 검토" in o:
            return "사람 검토"
        return "기타"
    c4 = Counter(v4b(r["outcome"]) for r in v4_rows)
    _c(ws, rr, 1, "v4(물렁) → v5(확정 탈락) 대비", bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    rr += 1
    _hdr(ws, rr, ["지표", "v4", "v5"], [30, 10, 10])
    rr += 1
    for lab, a, b in [("확정 탈락(v4는 스케일업 안내로 유보)", c4.get("스케일업 안내", 0),
                       c5.get("확정 탈락", 0)),
                      ("사람 검토", c4.get("사람 검토", 0), c5.get("사람 검토 (경계)", 0))]:
        _c(ws, rr, 1, lab, size=9)
        _c(ws, rr, 2, a, align="center")
        _c(ws, rr, 3, b, align="center", bold=True)
        rr += 1

    # 시트2 확정 탈락 목록 (사유 포함)
    ws2 = wb.create_sheet("확정_탈락")
    ws2.sheet_view.showGridLines = False
    _c(ws2, 1, 1, f"확정 탈락 {len(fails):,}개사 — 사유별 (운영자가 큐를 비울 수 있게)",
       bold=True, size=12)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _hdr(ws2, 3, ["국문명", "업종(CB)", "스테이지", "→ 트랙", "탈락 사유"],
         [22, 20, 12, 10, 50])
    rr = 4
    for r in sorted(fails, key=lambda x: (x["track"], x["name_ko"])):
        _c(ws2, rr, 1, r["name_ko"], size=9, fill=V5_FILL["확정 탈락"])
        _c(ws2, rr, 2, r["sector"][:30], size=8, fill=V5_FILL["확정 탈락"])
        _c(ws2, rr, 3, r["stage"], size=8, align="center", fill=V5_FILL["확정 탈락"])
        _c(ws2, rr, 4, r["track"], size=9, align="center", fill=V5_FILL["확정 탈락"])
        _c(ws2, rr, 5, r.get("reasons", ""), size=8, fill=V5_FILL["확정 탈락"])
        rr += 1
    ws2.freeze_panes = "A4"

    # 시트3 전체 판정
    ws3 = wb.create_sheet("전체_판정_v5")
    ws3.sheet_view.showGridLines = False
    _c(ws3, 1, 1, f"전체 판정표 v5 ({len(v5_rows):,}개사)", bold=True, size=12)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _hdr(ws3, 3, ["국문명", "업종(CB)", "스테이지", "재단분류", "→ 트랙", "밴드",
                  "v5 판정", "사유"], [20, 18, 11, 16, 9, 9, 22, 30])
    rr = 4
    ordr = {b: i for i, b in enumerate(V5_ORDER)}
    for r in sorted(v5_rows, key=lambda x: (ordr.get(x["outcome"], 9), x["track"])):
        f = V5_FILL.get(r["outcome"], GRY)
        _c(ws3, rr, 1, r["name_ko"], fill=f, size=9)
        _c(ws3, rr, 2, r["sector"][:34], fill=f, size=8)
        _c(ws3, rr, 3, r["stage"], fill=f, size=8, align="center")
        _c(ws3, rr, 4, r["type"][:24], fill=f, size=8)
        _c(ws3, rr, 5, r["track"], fill=f, size=9, align="center")
        _c(ws3, rr, 6, r["band"], fill=f, size=8, align="center")
        _c(ws3, rr, 7, r["outcome"], fill=f, size=8)
        _c(ws3, rr, 8, (r.get("reasons") or "")[:40], fill=f, size=8)
        rr += 1
    ws3.freeze_panes = "A4"

    # 시트4 방법론
    ws4 = wb.create_sheet("방법론_v5")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 115
    notes = [
        ("엔진 v5 — '확실히 아닌 건 확실히 탈락'", True, 12),
        ("확정 탈락 디스퀄리파이어 (프로그램이 절대 waive 안 하는 기준)", True, 11),
        ("① 스테이지 이탈: 시리즈B+ (양 트랙), HAX 는 시리즈A도 이탈(프리시드~시드 전용). "
         "② 섹터 부적합: HAX 제외 섹터(핀테크·크립토·보안·이커머스·순수SW). "
         "③ 언어: C레벨 영어 불가 확인. ④ 제품: 동작 프로토타입 없음 확인. "
         "⑤ 커밋: 풀타임/이주 거부 확인(500). ⑥ HAX 오너십: 프라이스드 라운드/10% 불가.",
         False, 10),
        ("확인된 것만 탈락 — 추측으로 떨구지 않는다", True, 11),
        ("③④⑤⑥은 설문·덱이 있어야 확인된다. DB 대규모 단계에는 그 입력이 없어 "
         "**스테이지·섹터만 확정 탈락**을 발동하고, 언어·제품·커밋은 조건부(설문 필요)로 "
         "남긴다. 실제 지원서(설문+덱)가 들어오면 나머지 디스퀄리파이어가 활성화된다.",
         False, 10),
        ("v4에서 무엇을 되돌렸나", True, 11),
        ("v4의 '스케일업 트랙 안내'(물렁한 유보)를 철회했다. 시리즈B+·HAX 시리즈A는 "
         "프로그램에 명백히 부적합하므로 스케일업 안내가 아니라 **확정 탈락(스테이지 "
         "이탈)**으로 떨군다 — 운영자가 큐를 확실히 비울 수 있게. 사유가 붙어 있어 "
         "왜 탈락인지 즉시 판단 가능.", False, 10),
        ("불변·한계", True, 11),
        ("WEIGHTS·컷오프 불변. 확정 탈락은 '확인된 부적합 사실'에만 발동(오탈락 방지). "
         "키워드 라우터 근사 → 라우팅 불안정은 별도 사람 확인. 4축 점수는 팩트시트 있는 "
         "102개사에만.", False, 10),
    ]
    for i, (t, b, sz) in enumerate(notes, 1):
        c = _c(ws4, i, 1, t, bold=b, size=sz, wrap=True)
        if b and sz >= 11:
            c.fill = SUB
        ws4.row_dimensions[i].height = 48 if len(t) > 70 else 16

    OUT_V5.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_V5)
    return OUT_V5


OUT_V6 = OUT.parent / "gbd_auto_eval_v7.xlsx"

RED = PatternFill("solid", fgColor="F4B7B7")
BLUE = PatternFill("solid", fgColor="DEEBF7")


def _v6_fill(zone):
    from screening import disqualifiers as dq
    if zone == dq.Z_FAIL:
        return RED
    if zone == dq.Z_BIO:
        return BLUE
    if "입력 없음" in zone:
        return GRY
    return GREEN


# 판정 → 액션 그룹(정렬·표시). 사람검토 폐지 → 메일 대상 vs 확정 탈락 2갈래.
def action_group(zone):
    from screening import disqualifiers as dq
    if zone == dq.Z_FAIL:
        return 3, "🔴 확정 탈락"
    if zone == dq.Z_BIO:
        return 2, "🔵 IndieBio 리퍼럴"
    if "입력 없음" in zone:
        return 4, "⚪ 대상외 (정보 공란)"
    return 1, "🟢 메일 대상"


def target_market(target):
    """타겟 국가 필드 → (표시 라벨, 미국 정합 여부)."""
    t = (target or "").strip()
    if not t:
        return "미상 (설문)", False
    return t, ("미국" in t)


def _engine_sheet(wb, title, header_lines, rows, track):
    """500/HAX 엔진 시트 — 업종/분야 분리 · 액션 2갈래 정렬 · 메일 유형 · 유사 합격사."""
    from screening import similar_admits, gate_v4
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for i, line in enumerate(header_lines):
        _c(ws, i + 1, 1, line, bold=(i == 0), size=12 if i == 0 else 9, wrap=True)
        ws.merge_cells(start_row=i + 1, start_column=1, end_row=i + 1, end_column=11)
        if i > 0:
            ws.row_dimensions[i + 1].height = 26
    from collections import Counter
    gc = Counter(action_group(r["outcome"])[1] for r in rows)
    hr = len(header_lines) + 1
    _c(ws, hr, 1, "액션 요약:  " + "   ".join(
        f"{g} {n}" for g, n in sorted(gc.items())), bold=True, size=9, wrap=True)
    ws.merge_cells(start_row=hr, start_column=1, end_row=hr, end_column=11)
    ws.row_dimensions[hr].height = 24
    hr += 1
    _hdr(ws, hr, ["액션", "국문명", "업종(CB)", "분야", "스테이지", "타겟 시장",
                  "판정", "사유", "메일 유형", f"유사 {track.upper()} 합격사", "재단분류"],
         [18, 17, 14, 15, 11, 12, 14, 24, 18, 24, 12])
    rr = hr + 1
    rows_sorted = sorted(rows, key=lambda r: (action_group(r["outcome"])[0],
                                              r["name_ko"]))
    for r in rows_sorted:
        z = r["outcome"]
        f = _v6_fill(z)
        band = gate_v4.band_of(r["stage"])
        sim = similar_admits.match_str(track, r.get("field", ""), r["desc"], band)
        tm, us = target_market(r.get("target", ""))
        field = ("⚠ " if r.get("mismatch") else "") + r.get("field", "미분류")
        _c(ws, rr, 1, action_group(z)[1], fill=f, size=8, bold=True)
        _c(ws, rr, 2, r["name_ko"], fill=f, size=9)
        _c(ws, rr, 3, r.get("cb_group", "—"), fill=f, size=8)
        _c(ws, rr, 4, field, fill=f, size=8, bold=True)
        _c(ws, rr, 5, r["stage"], fill=f, size=8, align="center")
        _c(ws, rr, 6, ("🇺🇸 " if us else "") + tm, fill=f, size=8, align="center")
        _c(ws, rr, 7, z, fill=f, size=8)
        _c(ws, rr, 8, (r.get("reasons") or "")[:44], fill=f, size=8)
        _c(ws, rr, 9, r.get("email", "—"), fill=f, size=8)
        _c(ws, rr, 10, sim[:42], fill=f, size=8)
        _c(ws, rr, 11, r["type"][:18], fill=f, size=8)
        rr += 1
    ws.freeze_panes = "A%d" % (hr + 1)


def build_v6(v6_rows) -> Path:
    from collections import Counter
    from screening import disqualifiers as dq
    wb = Workbook()

    # 시트1 요약
    ws = wb.active
    ws.title = "요약_v7"
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, f"디캠프 GBD DB Ver.26.01 — 엔진 v7 (분야 기반·사람검토 폐지) "
       f"{len(v6_rows):,}개사", bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    _c(ws, 2, 1, "업종(CB 그룹)은 거칠어 참고만 하고, 라우팅·탈락은 **분야**(사업 실체)로 "
       "판정. 확정 탈락(스테이지 이탈·분야 부적합·확인된 언어/제품)이 아니면 전부 "
       "**메일 대상**으로 흡수(사람검토 버킷 폐지 — 애매해도 메일은 보낸다). "
       "500/HAX 별개 시트, 접전은 양쪽 평가, HAX 탈락은 500 리퍼럴.", size=9, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.row_dimensions[2].height = 52
    c = Counter(r["outcome"] for r in v6_rows)
    _hdr(ws, 4, ["판정", "기업 수"], [40, 12])
    rr = 5
    for k, n in c.most_common():
        _c(ws, rr, 1, k, fill=_v6_fill(k))
        _c(ws, rr, 2, n, align="center", bold=True, fill=_v6_fill(k))
        rr += 1
    rr += 1
    mail = sum(n for k, n in c.items() if k not in (dq.Z_FAIL, dq.Z_OOS))
    _c(ws, rr, 1, f"→ 🟢 메일 대상 합계(IndieBio 포함): {mail:,}개사  /  "
       f"🔴 확정 탈락: {c.get(dq.Z_FAIL, 0):,}개사  /  "
       f"업종≠분야 불일치 신호: {sum(1 for r in v6_rows if r.get('mismatch')):,}개사  /  "
       f"크로스 리퍼럴(HAX→500): {sum(1 for r in v6_rows if r.get('cross')):,}개사",
       bold=True, wrap=True)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
    ws.row_dimensions[rr].height = 30

    # 시트2 500 엔진
    s500 = [r for r in v6_rows if r["primary"] in ("500", "500/hax")
            or r.get("cross") == "500"]
    _engine_sheet(wb, "500_엔진", [
        f"500 Global Flagship 엔진 — {len(s500):,}개사",
        "축: 트랙션40·팀30·시장20·해자10 / 대상: MVP+트랙션, 섹터 무관, 영어 전용·SV 상주",
        "트랙션=유료고객·매출(MRR/ARR)·성장률·리텐션(B2B는 계약/PO/LOI). 시리즈B+ 확정 탈락.",
    ], s500, "500")

    # 시트3 HAX 엔진
    shax = [r for r in v6_rows if r["primary"] in ("hax", "500/hax")]
    _engine_sheet(wb, "HAX_엔진", [
        f"HAX (SOSV) 엔진 — {len(shax):,}개사",
        "축: TRL40·팀30·양산20·고객10 / 대상: 프리시드~시드 하드테크(기후·로보틱스·소재·헬스HW)",
        "제외 분야(핀테크·크립토·커머스·보안)·시리즈A+ = 확정 탈락. 캡 없는 SAFE+지분10%.",
    ], shax, "hax")

    # 시트4 확정 탈락 (사유)
    fails = [r for r in v6_rows if r["outcome"] == dq.Z_FAIL]
    ws4 = wb.create_sheet("확정_탈락")
    ws4.sheet_view.showGridLines = False
    _c(ws4, 1, 1, f"확정 탈락 {len(fails):,}개사 — 사유 명시 (운영자가 큐를 비울 수 있게)",
       bold=True, size=12)
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _hdr(ws4, 3, ["국문명", "업종(CB)", "분야", "스테이지", "대상 트랙", "탈락 사유"],
         [22, 16, 15, 12, 10, 46])
    rr = 4
    for r in sorted(fails, key=lambda x: (str(x["primary"]), x["name_ko"])):
        _c(ws4, rr, 1, r["name_ko"], size=9, fill=RED)
        _c(ws4, rr, 2, r.get("cb_group", "—"), size=8, fill=RED)
        _c(ws4, rr, 3, r.get("field", ""), size=8, fill=RED)
        _c(ws4, rr, 4, r["stage"], size=8, align="center", fill=RED)
        _c(ws4, rr, 5, str(r["primary"]), size=8, align="center", fill=RED)
        _c(ws4, rr, 6, r.get("reasons", ""), size=8, fill=RED)
        rr += 1
    ws4.freeze_panes = "A4"

    # 시트5 방법론
    ws5 = wb.create_sheet("방법론_v7")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 115
    notes = [
        ("엔진 v7 — 업종/분야 분리 · 사람검토 폐지 · 새 DB(Ver.26.01)", True, 12),
        ("업종 ≠ 분야", True, 11),
        ("업종(CB 40그룹)은 거칠다 — 'Hardware' 그룹에 로보틱스·반도체와 소매기술·화상회의가, "
         "'Financial Services'에 결제·대출·보험이 섞인다. 그래서 라우팅·탈락은 **분야**"
         "(1줄 소개·기술로 세밀 판정)를 기준으로 하고, 업종은 거친 prior/폴백으로만 쓴다. "
         "업종≠분야 불일치(⚠)는 그 자체가 검토 신호.", False, 10),
        ("사람검토 폐지 → 경계는 메일", True, 11),
        ("확정 탈락이 아닌 모든 것은 '메일 대상'으로 흡수한다. 애매하다고 붙잡아두면 "
         "아무 일도 안 일어난다 — 경계·설문 필요·점수화 가능·라우팅 접전 전부 메일로. "
         "메일 유형(내부검토·지원안내 / 설문·자료요청 / 자가진단·보완안내)만 세분한다.", False, 10),
        ("확정 탈락 = 확인된 부적합만", True, 11),
        ("스테이지 이탈(시리즈B+ 양 트랙·HAX 시리즈A+), 분야 부적합(HAX 제외 분야), "
         "확인된 언어불가·제품없음. DB로 확인 가능한 스테이지·분야가 주력이고, 언어·제품은 "
         "덱/설문이 와야 확인되므로 그전엔 메일(설문)로 남긴다.", False, 10),
        ("새 DB(Ver.26.01) 재추출", True, 11),
        ("이전 판(gbd_full.json)은 빈 행이 42% 섞여 '대상외'가 과다했다 — 추출 문제였다. "
         "새 파일은 업종·기술·스테이지·소개가 99% 채워져 대상외가 사라지고, 스테이지 이탈 "
         "확정 탈락이 제대로 발동한다. 타겟 국가는 1%만 채워져 미국/일본 구분은 설문 과제로 남음.",
         False, 10),
        ("트랙션 정의(500)", True, 11),
        ("트랙션 = 시장이 실제 반응한 증거: 유료고객 수·매출(MRR/ARR)·MoM 성장률·리텐션, "
         "B2B는 계약/PO/LOI. 1줄 소개론 측정 불가 → 대규모 자동단계엔 트랙션 점수를 내지 "
         "않고, 덱·설문 있는 기업만 채점한다.", False, 10),
        ("불변·한계", True, 11),
        ("WEIGHTS·컷오프 불변. 확정 탈락은 확인된 부적합만. 분야 분류·유사 매칭은 근사. "
         "4축 점수는 팩트시트 있는 소표본에만.", False, 10),
    ]
    for i, (t, b, sz) in enumerate(notes, 1):
        cc = _c(ws5, i, 1, t, bold=b, size=sz, wrap=True)
        if b and sz >= 11:
            cc.fill = SUB
        ws5.row_dimensions[i].height = 46 if len(t) > 70 else 16

    OUT_V6.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_V6)
    return OUT_V6


OUT_UF = OUT.parent / "us_forged_candidates.xlsx"


def build_us_forged(uf_rows) -> Path:
    """US FORGED(디캠프 x HAX Hardtech Pre-Program) 발송 후보 워크북."""
    from collections import Counter
    from screening import similar_admits, gate_v4
    wb = Workbook()
    elig = [r for r in uf_rows if r["uf_status"].startswith("적합")]
    seed = [r for r in elig if r["uf_stage"] == "OK"]
    unk = [r for r in elig if r["uf_stage"] == "UNKNOWN"]
    us = [r for r in elig if r["uf_us"]]
    nonhard = sum(1 for r in uf_rows
                  if r["uf_status"] == "부적합" and "분야" in r["uf_reasons"])
    hard_late = sum(1 for r in uf_rows
                    if r["uf_status"] == "부적합" and "스테이지" in r["uf_reasons"])
    hard_total = hard_late + len(elig)
    T = Counter(r["uf_tier"] for r in elig)

    # 시트1 요약 (빼기 → 남은 수 형식)
    ws = wb.active
    ws.title = "요약_US_FORGED"
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, "US FORGED — 디캠프 x HAX Hardtech Pre-Program 발송 후보 (공고문 요건)",
       bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _c(ws, 2, 1, "요건: 미국 진출 준비 Pre-Seed~Seed 딥테크·하드테크. Software-only·"
       "일반 소비재·범용 제품 제외. Lab-scale 이상 프로토타입. 선발 8~10개사(마감 9/6). "
       "→ 각 줄은 그 필터까지 적용하고 '남은' 기업 수(누적).", size=9, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 40
    _hdr(ws, 4, ["필터 단계 (빼기 → 남은 수)", "제외", "남은 수"], [42, 10, 12])
    funnel = [("전체 DB", "", len(uf_rows), None),
              ("  └ Software-only·소비재 제거", f"-{nonhard}", None, None),
              ("① 하드테크 분야 (남은 수)", "", hard_total, None),
              ("  └ 시리즈A+ (스테이지 이탈) 제거", f"-{hard_late}", None, None),
              ("② Pre-Seed~Seed/미상 (남은 수 = 발송 후보)", "", len(elig), GREEN),
              ("     ├ 스테이지 시드 확정", "", len(seed), None),
              ("     └ 스테이지 미상 (설문 먼저)", "", len(unk), None),
              ("③ + 미국 진출 명시 (DB 기재분)", "", len(us), None)]
    rr = 5
    for lab, minus, n, fill in funnel:
        _c(ws, rr, 1, lab, fill=fill)
        _c(ws, rr, 2, minus, align="center", fill=fill)
        _c(ws, rr, 3, "" if n is None else n, align="center", bold=True, fill=fill)
        rr += 1
    rr += 1
    _c(ws, rr, 1, "신뢰도 티어 (발송 후보 안에서 — 더 세게 거르는 대신 티어로 분류)",
       bold=True, size=11, fill=SUB)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    rr += 1
    _hdr(ws, rr, ["티어", "조건", "수"], [16, 40, 10])
    rr += 1
    tiers = [("T1 최우선", "시드 확정 + 분야 소개로 확정", GREEN),
             ("T2 검토", "시드확정+분야폴백 / 미상+분야확정 (둘 중 하나만)", YEL),
             ("T3 설문 우선", "스테이지 미상 + 분야 폴백 (약한 신호)", None)]
    for name, cond, fill in tiers:
        _c(ws, rr, 1, name, fill=fill, bold=True)
        _c(ws, rr, 2, cond, fill=fill)
        _c(ws, rr, 3, T.get(name, 0), align="center", bold=True, fill=fill)
        rr += 1
    rr += 1
    _c(ws, rr, 1, "※ DB만으로 더 세게 거르면 진짜 딥테크(특수 용어라 분야폴백)까지 날아간다. "
       "T1 아래는 '더 거르기'가 아니라 '설문으로 확인'이 맞다. US FORGED 핵심 요건"
       "(Lab-scale 프로토타입·미국 의지·기술 차별성)은 DB에 없어 설문/덱 필수.",
       size=9, wrap=True)
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    ws.row_dimensions[rr].height = 40

    # 시트2 발송 후보 (티어 정렬)
    ws2 = wb.create_sheet("발송_후보")
    ws2.sheet_view.showGridLines = False
    _c(ws2, 1, 1, f"US FORGED 발송 후보 {len(elig):,}개사 — 티어순(T1 최우선 상단)",
       bold=True, size=12)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    _hdr(ws2, 3, ["티어", "국문명", "분야", "스테이지", "타겟 시장", "설문/확인 항목",
                  "업종(CB)", "유사 HAX 합격사", "재단분류"],
         [14, 18, 14, 11, 12, 24, 14, 24, 12])
    torder = {"T1 최우선": 0, "T2 검토": 1, "T3 설문 우선": 2}
    tfill = {"T1 최우선": GREEN, "T2 검토": YEL, "T3 설문 우선": None}
    rows_sorted = sorted(elig, key=lambda r: (torder.get(r["uf_tier"], 3),
                                              r["uf_field"], r["name_ko"]))
    rr = 4
    for r in rows_sorted:
        f = tfill.get(r["uf_tier"])
        band = gate_v4.band_of(r["stage"])
        sim = similar_admits.match_str("hax", r["uf_field"], r["desc"], band)
        tm = ("🇺🇸 " if r["uf_us"] else "") + ((r.get("target") or "").strip() or "미상")
        _c(ws2, rr, 1, r["uf_tier"], fill=f, size=8, bold=True)
        _c(ws2, rr, 2, r["name_ko"], fill=f, size=9)
        _c(ws2, rr, 3, r["uf_field"], fill=f, size=8, bold=True)
        _c(ws2, rr, 4, r["stage"], fill=f, size=8, align="center")
        _c(ws2, rr, 5, tm, fill=f, size=8, align="center")
        _c(ws2, rr, 6, r["uf_reasons"][:38], fill=f, size=8)
        _c(ws2, rr, 7, r["sector"][:20], fill=f, size=8)
        _c(ws2, rr, 8, sim[:42], fill=f, size=8)
        _c(ws2, rr, 9, r["type"][:18], fill=f, size=8)
        rr += 1
    ws2.freeze_panes = "A4"

    # 시트3 부적합 사유
    fails = [r for r in uf_rows if r["uf_status"] == "부적합"]
    ws3 = wb.create_sheet("부적합_참고")
    ws3.sheet_view.showGridLines = False
    _c(ws3, 1, 1, f"부적합 {len(fails):,}개사 — 사유 분포 (참고)", bold=True, size=12)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    rc = Counter("Software-only/소비재(하드테크 아님)" if "분야" in r["uf_reasons"]
                 else "스테이지 이탈(시리즈A+)" for r in fails)
    _hdr(ws3, 3, ["부적합 사유", "기업 수"], [40, 12])
    rr = 4
    for k, n in rc.most_common():
        _c(ws3, rr, 1, k)
        _c(ws3, rr, 2, n, align="center", bold=True)
        rr += 1

    OUT_UF.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_UF)
    return OUT_UF


if __name__ == "__main__":
    from screening import gbd_pipeline
    rows = gbd_pipeline.run()
    print(build(rows, gbd_pipeline.summarize(rows)))
