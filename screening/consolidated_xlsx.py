"""통합 평가 워크북 — 전체 기업 한 파일. output/screening/engine_full_eval.xlsx

  python -m screening.consolidated_xlsx

시트
  1) 전체 평가   — 모든 기업 한 줄씩 (그룹·트랙·밴드·게이트·v2·v3·비고)
  2) 퍼널 요약   — 그룹×1차필터 교차표 + 카운트
  3) 축별 레벨   — 점수화된 기업 4축 레벨·근거
  4) 방법론·한계
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from screening import consolidated, dataset, rules
from screening.live_batch import ROUTING as ROUTING3

BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "output" / "screening" / "engine_full_eval.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
GRP_FILL = PatternFill("solid", fgColor="D6DCE4")
GREEN = PatternFill("solid", fgColor="C6EFCE")    # 확정 추천 / 합격
LGREEN = PatternFill("solid", fgColor="E2EFDA")   # 통과·점수화
YELLOW = PatternFill("solid", fgColor="FFF2CC")   # 사람 검토
ORANGE = PatternFill("solid", fgColor="FCE4D6")   # 탈락·라우팅
RED = PatternFill("solid", fgColor="F8CBAD")      # 확정 비추천
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _c(ws, r, c, v, *, bold=False, fill=None, wrap=True, align="left", size=10):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=FONT, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = fill
    return cell


def _hdr(ws, r, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = _c(ws, r, i, h, bold=True, fill=HDR_FILL, align="center")
        c.font = HDR_FONT
        ws.column_dimensions[get_column_letter(i)].width = w


def _row_fill(row) -> PatternFill:
    if row["gate"] in ("라우팅", rules.GATE_FAIL):
        return ORANGE
    if row["v3"] == rules_zone("확정 비추천"):
        return RED
    if row["v3"] == "확정 추천":
        return GREEN
    if row["gate"] == rules.GATE_HUMAN or row["v3"] == "사람 검토":
        return YELLOW
    return LGREEN


def rules_zone(z):  # 상수 오타 방지용 헬퍼
    return z


def build() -> Path:
    rows = consolidated.all_rows()
    wb = Workbook()

    # ---------------- 시트 1: 전체 평가 ----------------
    ws = wb.active
    ws.title = "전체_평가"
    ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, f"500/HAX 프리스크리닝 엔진 — 전체 평가 {len(rows)}개사",
       bold=True, size=13, wrap=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    _c(ws, 2, 1, "카드몬스터·올세일(실제 500 선발) 포함. 웹 검색 수집(직접 크롤링 "
       "차단) → 전 기업 간이 진단. 1차 필터(트랙 라우팅+하드 게이트)가 스테이지·"
       "섹터를 먼저 거르고, 통과한 기업만 v2 점추정·v3 구간추정으로 점수화한다.",
       size=9, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws.row_dimensions[2].height = 40

    _hdr(ws, 4, ["그룹", "기업", "라벨/배치", "트랙", "밴드", "① 게이트",
                 "② v2 Tier(점추정)", "③ v3 구역(구간)", "비고"],
         [22, 22, 14, 6, 10, 10, 19, 17, 26])
    r = 5
    for row in rows:
        fill = _row_fill(row)
        v2 = row["v2"] if row["v2w"] is None else f"{row['v2']} ({row['v2w']:.2f})"
        v3 = (row["v3"] if row["v3lo"] is None
              else f"[{row['v3lo']:.2f},{row['v3hi']:.2f}] {row['v3']}")
        _c(ws, r, 1, row["group"], size=8, fill=fill)
        _c(ws, r, 2, row["name"], fill=fill)
        _c(ws, r, 3, row["tag"], size=8, align="center", fill=fill)
        _c(ws, r, 4, row["track"], align="center", fill=fill)
        _c(ws, r, 5, row["band"], align="center", size=9, fill=fill)
        _c(ws, r, 6, row["gate"], align="center", size=9, fill=fill)
        _c(ws, r, 7, v2, size=9, fill=fill)
        _c(ws, r, 8, v3, size=9, fill=fill)
        _c(ws, r, 9, row["note"], size=8, fill=fill)
        r += 1
    ws.freeze_panes = "A5"
    r += 1
    _c(ws, r, 1, "범례:", bold=True, wrap=False)
    for i, (lab, f) in enumerate([("확정추천/합격", GREEN), ("통과·점수화", LGREEN),
                                  ("사람검토", YELLOW), ("탈락/라우팅", ORANGE),
                                  ("확정비추천", RED)]):
        _c(ws, r, 2 + i, lab, fill=f, wrap=False, size=9, align="center")

    # ---------------- 시트 2: 퍼널 요약 ----------------
    ws2 = wb.create_sheet("퍼널_요약")
    ws2.sheet_view.showGridLines = False
    _c(ws2, 1, 1, "1차 필터 퍼널 — 그룹 × 게이트 결과", bold=True, size=12, wrap=False)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    groups = ["G1 실전 500 선발(딥크롤)", "G2 라벨·대조 표본",
              "G3 디캠프 배치 2·4·6·7기", "G4 디캠프 배치 1·3·5기",
              "G5 실제 500/HAX 포트폴리오"]
    gate_cats = ["통과/조건부(점수화)", "사람 검토", "게이트 탈락", "라우팅"]

    def cat(row):
        if row["gate"] == "라우팅":
            return "라우팅"
        if row["gate"] == rules.GATE_FAIL:
            return "게이트 탈락"
        if row["gate"] == rules.GATE_HUMAN:
            return "사람 검토"
        return "통과/조건부(점수화)"

    cross = {g: Counter() for g in groups}
    for row in rows:
        if row["group"] in cross:
            cross[row["group"]][cat(row)] += 1
    _hdr(ws2, 3, ["그룹"] + gate_cats + ["합계"], [28, 18, 12, 12, 10, 8])
    rr = 4
    for g in groups:
        if sum(cross[g].values()) == 0:
            continue
        _c(ws2, rr, 1, g, size=9)
        tot = 0
        for j, gc in enumerate(gate_cats, 2):
            n = cross[g][gc]
            tot += n
            _c(ws2, rr, j, n or "", align="center",
               fill=ORANGE if gc in ("게이트 탈락", "라우팅") and n else
               YELLOW if gc == "사람 검토" and n else
               LGREEN if gc.startswith("통과") and n else None)
        _c(ws2, rr, 6, tot, align="center", bold=True)
        rr += 1
    _c(ws2, rr, 1, "합계", bold=True, fill=GRP_FILL)
    for j, gc in enumerate(gate_cats, 2):
        _c(ws2, rr, j, sum(cross[g][gc] for g in groups), align="center",
           bold=True, fill=GRP_FILL)
    _c(ws2, rr, 6, len(rows), align="center", bold=True, fill=GRP_FILL)

    rr += 3
    _c(ws2, rr, 1, "점수화된 기업의 v3 구역 분포 — 자동 보강 전 → 후", bold=True,
       size=11, wrap=False)
    rr += 1
    _c(ws2, rr, 1, "자동 보강 = 창업자 이력·트랙션을 추가 크롤링해 `확인 필요` Team 축을 "
       "실증으로 채우는 단계(Step 1.5). 사람 검토를 줄이는 유일한 정당한 방법은 규칙 "
       "변경이 아니라 데이터 보강이다(덱 부재로 인한 Market null 은 여전히 남는다).",
       size=8, wrap=True)
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    ws2.row_dimensions[rr].height = 30
    rr += 1
    before = consolidated.all_rows(enrich=False)
    zb = Counter(row["v3"] for row in before
                 if row["v3"] in ("확정 추천", "사람 검토", "확정 비추천"))
    zc = Counter(row["v3"] for row in rows
                 if row["v3"] in ("확정 추천", "사람 검토", "확정 비추천"))
    _hdr(ws2, rr, ["v3 구역", "보강 전", "보강 후"], [20, 10, 10])
    rr += 1
    for z in ("확정 추천", "사람 검토", "확정 비추천"):
        _c(ws2, rr, 1, z, fill=GREEN if z == "확정 추천" else
           YELLOW if z == "사람 검토" else RED)
        _c(ws2, rr, 2, zb.get(z, 0), align="center")
        _c(ws2, rr, 3, zc.get(z, 0), align="center", bold=True)
        rr += 1

    # ---------------- 시트 3: 축별 레벨 ----------------
    ws3 = wb.create_sheet("축별_레벨")
    ws3.sheet_view.showGridLines = False
    _c(ws3, 1, 1, "축별 레벨 분류 — 점수화된 기업 (근거 포함)",
       bold=True, size=12, wrap=False)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _hdr(ws3, 3, ["기업", "트랙", "주축(Traction/TRL)", "Team",
                  "Market/양산", "Moat/고객"], [20, 6, 30, 24, 24, 24])
    AX = {"500": ("traction", "team", "market", "moat"),
          "hax": ("trl", "team", "manufacturing", "customer")}
    rr = 4
    rr = _axis_block(ws3, rr, AX)
    ws3.freeze_panes = "A4"

    # ---------------- 시트 4: 방법론 ----------------
    _methodology(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


def _axis_levels():
    """{key: (track, {axis:(lv,why)})} — 모든 출처의 레벨을 모은다."""
    out = {}
    # G1: enriched
    from screening import live_eval
    for k in consolidated.G1_KEYS:
        e = live_eval.ENRICHED[k]
        out[k] = (e["track"], {a: (v[0], v[1]) for a, v in e["levels"].items()})
    # G2: dataset Fable blind (개선 후) 우선, 없으면 LEVELS_V2
    from screening import agreement
    fable = agreement.merged_fable(True)
    for c in dataset.COMPANIES:
        if c.key in consolidated.G1_KEYS or c.track == "bio_routing":
            continue
        lv = fable.get(c.key) or dataset.LEVELS_V2.get(c.key, {})
        out[c.key] = (c.track, {a: (v[0], v[1]) for a, v in lv.items()})
    # G3
    from screening import levels_live
    for k, axes in levels_live.LEVELS_LIVE.items():
        out[k] = (ROUTING3[k][0], {a: (v[0], v[1]) for a, v in axes.items()})
    # G4 (자동 보강 오버레이 반영)
    overlay = consolidated._enrichment()
    try:
        from screening import levels_live2, live_batch2
        for k, axes in levels_live2.LEVELS_LIVE2.items():
            out[k] = (live_batch2.ROUTING2[k][0],
                      {a: (v[0], v[1]) for a, v in axes.items()})
    except Exception:
        pass
    # G3+G4 null 축에 보강분 덮어쓰기 (표시용)
    for k, axes in overlay.items():
        if k in out:
            track, merged = out[k]
            for a, ev in axes.items():
                if a in merged and merged[a][0] is None and ev[0] is not None:
                    merged[a] = (ev[0], ev[1])
    # G5
    try:
        from screening import levels_portfolio, live_portfolio
        pf = live_portfolio.load_facts()
        for k, axes in levels_portfolio.LEVELS_PORTFOLIO.items():
            out[k] = (pf[k]["program"], {a: (v[0], v[1]) for a, v in axes.items()})
    except Exception:
        pass
    return out


def _axis_block(ws, rr, AX):
    levels = _axis_levels()
    name_of = {}
    for c in dataset.COMPANIES:
        name_of[c.key] = c.name
    for row in consolidated.all_rows():
        name_of[row["key"]] = row["name"]
    for key, (track, axes) in levels.items():
        if track not in AX:
            continue
        _c(ws, rr, 1, name_of.get(key, key), bold=True, size=9)
        _c(ws, rr, 2, track, align="center")
        for i, axis in enumerate(AX[track], 3):
            v = axes.get(axis)
            if v is None:
                _c(ws, rr, i, "—")
                continue
            lv, why = v
            tag = f"L{lv}" if lv else "확인 필요"
            _c(ws, rr, i, f"{tag} — {why}", size=8)
        rr += 1
    return rr


def _methodology(wb):
    ws = wb.create_sheet("방법론_한계")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 115
    notes = [
        ("방법론 및 한계", True, 12),
        ("엔진 파이프라인", True, 11),
        ("Step3 트랙 라우팅(바이오→IndieBio / 순수SW→500 / 하드웨어→HAX) → "
         "Step4 하드 게이트(제외 섹터·스테이지·프로토타입) → 여기까지 통과한 기업만 "
         "Step6 축별 레벨 분류 → 가중평균 → v2 Tier / v3 구간 → Step7 Fit → 조치. "
         "가중치(Traction/TRL 40·Team 30·Market/양산 20·Moat/고객 10)는 불변.", False, 10),
        ("평가 대상 그룹 (총 102개사)", True, 11),
        ("G1 카드몬스터·올세일 = 실제 500 선발사(딥크롤 보강, 2). G2 = 기존 라벨·대조 "
         "표본 18개사(합격/탈락/미확인). G3 = 디캠프 배치 2·4·6·7기 31개사. "
         "G4 = 디캠프 배치 1·3·5기 24개사. G5 = 실제 500/HAX 포트폴리오 27개사"
         "(500 투자 15 + HAX 참여 12).", False, 10),
        ("자동 데이터 보강 (사람 검토 감축)", True, 11),
        ("사람 검토가 많은 주원인은 규칙이 아니라 데이터 부재다 — 개인 창업자 경력이 "
         "보도에 드물어 Team 축(가중치 0.30)이 `확인 필요`가 되면 v3 구간이 넓어져 "
         "추천선을 걸친다. 개선책으로 창업자 이력을 추가 크롤링해 실증이 확인된 Team "
         "축을 채웠다(못 찾으면 null 유지 — 지어내지 않음). 이는 규칙을 라벨에 맞추는 "
         "암기가 아니라 입력 데이터의 완결성을 높이는 자동화다. Market 축은 덱 부재로 "
         "규칙상 상향이 금지돼 대부분 null 로 남으며, 이는 실제 지원서에 덱이 붙으면 "
         "해소된다. 퍼널_요약 시트에 보강 전후 v3 분포를 병기했다.", False, 10),
        ("레벨 분류의 블라인드성", True, 11),
        ("G1·G3·G4 레벨은 dataset·정답을 본 적 없는 격리 세션이 개선된 §3·§4 규칙"
         "(판별 질문 포함)만 보고 분류. G2 는 블라인드 Fable 분류(개선 후) 사용. "
         "회사의 조달 실적을 Team 근거로 쓰지 않는 등 증거 등급 규칙 적용.", False, 10),
        ("데이터 수집 한계", True, 11),
        ("THE VC·dcamp.kr 직접 크롤링 차단(HTTP 403) → 웹 검색 결과 본문만 사용. "
         "증거 등급 '문서 명시'는 전부 언론·기업DB 표기, 피치덱·재무제표 없음 → "
         "전 기업 간이 진단. 스테이지 밴드는 공개 최신 라운드로 추정.", False, 10),
        ("이 평가로 말할 수 있는 것 / 없는 것", True, 11),
        ("● 말할 수 있는 것: 1차 필터가 스테이지·섹터를 실제로 거른다. 디캠프 배치는 "
         "500/HAX(프리시드~시드)보다 뒤 단계라 다수가 게이트 탈락하는 것이 구조적으로 "
         "예측된다. 카드몬스터·올세일(실제 500 선발)은 걸러지지 않고 추천 트랙에 오른다.",
         False, 10),
        ("● 말할 수 없는 것: 디캠프 배치사는 500/HAX 지원·합격 기업이 아니다 → "
         "정밀도·특이도 측정 불가. Team 축이 다수 '확인 필요'인 것은 개인 경력이 "
         "보도에 드물기 때문(엔진의 보수성)이며 실제 지원서엔 CV 가 붙으므로 완화된다.",
         False, 10),
        ("● G5 주의: 500 포트폴리오 15개사는 대부분 2013~2020 투자로 지금은 A 이후 "
         "단계다. 스테이지가 '지원 시점'이 아니라 '현재'라 사람 검토가 많이 나오는데, "
         "이는 '지금 지원하면 늦다'는 뜻이지 500 이 틀렸다는 뜻이 아니다. HAX "
         "포트폴리오 12개사는 프리시드~시드가 많아 스테이지 게이트를 잘 통과한다.",
         False, 10),
    ]
    for i, (txt, bold, size) in enumerate(notes, 1):
        c = _c(ws, i, 1, txt, bold=bold, size=size, wrap=True)
        if bold and size >= 11:
            c.fill = GRP_FILL
        ws.row_dimensions[i].height = 42 if len(txt) > 70 else 16


if __name__ == "__main__":
    print(build())
