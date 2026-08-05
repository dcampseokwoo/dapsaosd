"""디캠프 배치 31개사 평가 결과 → 엑셀 워크북 생성.

  python -m screening.live_batch_xlsx

시트 구성
  1) 요약 — 퍼널(라우팅→게이트→점수)과 최종 판정 한 줄씩
  2) 게이트 상세 — 트랙 라우팅 사유 + 스테이지 밴드 + 게이트 판정
  3) 축별 레벨 — 점수화 대상의 4축 레벨·근거
  4) 방법론·한계

레벨은 levels_live.py (격리 세션 분류)에서 읽는다.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from screening import rules, rules_v2, rules_v3
from screening.live_batch import ROUTING, funnel, load_facts

BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "output" / "screening" / "live_eval.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D6DCE4")
PASS_FILL = PatternFill("solid", fgColor="E2EFDA")   # 초록 — 점수화 진행
FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")   # 주황 — 탈락/라우팅
HUMAN_FILL = PatternFill("solid", fgColor="FFF2CC")  # 노랑 — 사람 검토
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _levels():
    try:
        from screening.levels_live import LEVELS_LIVE
        return LEVELS_LIVE
    except Exception:
        return {}


def _cell(ws, r, c, v, *, bold=False, fill=None, wrap=True, align="left", size=10):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=FONT, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = fill
    return cell


def _hdr(ws, r, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = _cell(ws, r, i, h, bold=True, fill=HDR_FILL, align="center")
        c.font = HDR_FONT
        ws.column_dimensions[get_column_letter(i)].width = w


def score_row(key: str, fr: dict) -> dict:
    """게이트 통과 기업의 v2/v3 점수. 레벨 없으면 None 필드."""
    lv = _levels().get(key)
    if not fr["scoreable"] or fr["track"] == "bio_routing" or lv is None:
        return {"v2_tier": "—", "v2_w": None, "v3_zone": "—", "v3_lo": None,
                "v3_hi": None, "levels": None}
    levels = {a: v[0] for a, v in lv.items()}
    unstable = {a: v[2] for a, v in lv.items() if len(v) > 2 and v[2] is not None}
    v2 = rules_v2.aggregate(fr["track"], levels)
    gate = rules.GATE_HUMAN if fr["gate"] == rules.GATE_HUMAN else rules.GATE_COND
    v3 = rules_v3.decide(fr["track"], levels, unstable, gate)
    return {"v2_tier": v2.tier, "v2_w": v2.weighted, "v3_zone": v3.zone,
            "v3_lo": v3.lo, "v3_hi": v3.hi, "levels": lv}


def build() -> Path:
    facts = load_facts()
    rows = funnel()
    lv_all = _levels()
    wb = Workbook()

    # ---------------- 시트 1: 요약 ----------------
    ws = wb.active
    ws.title = "요약"
    ws.sheet_view.showGridLines = False
    _cell(ws, 1, 1, "디캠프 배치 기업 500/HAX 프리스크리닝 엔진 평가 — 31개사",
          bold=True, size=13, fill=None, wrap=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _cell(ws, 2, 1, "배치 2·4·6·7기 실제 선발사. 웹 검색 수집(직접 크롤링 차단). "
          "디캠프 배치는 프리A~시리즈A 스케일업 프로그램이라 500/HAX(프리시드~시드) "
          "타깃보다 뒤 단계 → 1차 게이트에서 스테이지 탈락이 다수인 것이 정상.",
          size=9, wrap=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 42

    hdr = ["기업", "배치", "트랙", "밴드", "① 게이트", "② v2 Tier(점추정)",
           "③ v3 구역(구간)", "최종 판정"]
    _hdr(ws, 4, hdr, [22, 6, 7, 10, 14, 20, 16, 30])
    r = 5
    for fr in rows:
        s = score_row(fr["key"], fr)
        if fr["gate"] == rules.GATE_FAIL:
            fill = FAIL_FILL
        elif fr["gate"] == rules.GATE_HUMAN or fr["gate"] == "라우팅":
            fill = HUMAN_FILL if fr["gate"] == rules.GATE_HUMAN else FAIL_FILL
        else:
            fill = PASS_FILL
        v2 = s["v2_tier"] if s["v2_w"] is None else f"{s['v2_tier']} ({s['v2_w']:.2f})"
        v3 = s["v3_zone"] if s["v3_lo"] is None else \
            f"[{s['v3_lo']:.2f},{s['v3_hi']:.2f}] {s['v3_zone']}"
        _cell(ws, r, 1, fr["name"], fill=fill)
        _cell(ws, r, 2, fr["batch"], align="center", fill=fill)
        _cell(ws, r, 3, fr["track"], align="center", fill=fill)
        _cell(ws, r, 4, fr["band"], align="center", fill=fill)
        _cell(ws, r, 5, fr["gate"], align="center", fill=fill)
        _cell(ws, r, 6, v2, fill=fill)
        _cell(ws, r, 7, v3, fill=fill)
        _cell(ws, r, 8, fr["outcome"], fill=fill, size=9)
        r += 1
    # 범례
    r += 1
    _cell(ws, r, 1, "범례:", bold=True, wrap=False)
    _cell(ws, r, 2, "초록=점수화 진행", fill=PASS_FILL, wrap=False)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    _cell(ws, r, 4, "노랑=사람 검토(500 A이후)", fill=HUMAN_FILL, wrap=False)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    _cell(ws, r, 6, "주황=게이트 탈락/바이오 라우팅", fill=FAIL_FILL, wrap=False)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    ws.freeze_panes = "A5"

    # ---------------- 시트 2: 게이트 상세 ----------------
    ws2 = wb.create_sheet("게이트_라우팅")
    ws2.sheet_view.showGridLines = False
    _cell(ws2, 1, 1, "1차 필터 — 트랙 라우팅 + 하드 게이트 (점수화 이전)",
          bold=True, size=12, wrap=False)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _hdr(ws2, 3, ["기업", "배치", "트랙", "밴드", "라우팅·게이트 사유", "게이트 판정"],
         [22, 6, 7, 10, 60, 14])
    r = 4
    for fr in rows:
        fill = (FAIL_FILL if fr["gate"] in ("라우팅", rules.GATE_FAIL)
                else HUMAN_FILL if fr["gate"] == rules.GATE_HUMAN else PASS_FILL)
        _cell(ws2, r, 1, fr["name"], fill=fill)
        _cell(ws2, r, 2, fr["batch"], align="center", fill=fill)
        _cell(ws2, r, 3, fr["track"], align="center", fill=fill)
        _cell(ws2, r, 4, fr["band"], align="center", fill=fill)
        _cell(ws2, r, 5, fr["reason"], size=9, fill=fill)
        _cell(ws2, r, 6, fr["gate"], align="center", fill=fill)
        r += 1
    ws2.freeze_panes = "A4"

    # ---------------- 시트 3: 축별 레벨 ----------------
    ws3 = wb.create_sheet("축별_레벨")
    ws3.sheet_view.showGridLines = False
    _cell(ws3, 1, 1, "축별 레벨 분류 (게이트 통과 기업 — 격리 세션 분류, 개선 §3·§4)",
          bold=True, size=12, wrap=False)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _hdr(ws3, 3, ["기업", "트랙", "주축(Traction/TRL)", "Team",
                  "Market/양산", "Moat/고객"], [20, 7, 30, 26, 26, 26])
    r = 4
    AX = {"500": ("traction", "team", "market", "moat"),
          "hax": ("trl", "team", "manufacturing", "customer")}
    for fr in rows:
        lv = lv_all.get(fr["key"])
        if lv is None or fr["track"] == "bio_routing":
            continue
        _cell(ws3, r, 1, fr["name"], bold=True)
        _cell(ws3, r, 2, fr["track"], align="center")
        for i, axis in enumerate(AX[fr["track"]], 3):
            v = lv.get(axis)
            if v is None:
                _cell(ws3, r, i, "—")
                continue
            lvl, why = v[0], v[1]
            tag = f"L{lvl}" if lvl else "확인 필요"
            _cell(ws3, r, i, f"{tag} — {why}", size=8)
        r += 1
    ws3.freeze_panes = "A4"

    # ---------------- 시트 4: 방법론 ----------------
    ws4 = wb.create_sheet("방법론_한계")
    ws4.sheet_view.showGridLines = False
    notes = [
        ("방법론 및 한계", True, 12),
        ("", False, 10),
        ("표본", True, 11),
        ("디캠프 배치 2·4·6·7기에 실제 선발된 31개사. 배치 1·3·5기 및 미상 기업은 미포함.",
         False, 10),
        ("데이터 수집", True, 11),
        ("이 환경은 THE VC·dcamp.kr 등 직접 크롤링이 차단(HTTP 403)돼, 웹 검색 결과 "
         "본문만 사용했다. 따라서 증거 등급 '문서 명시'는 전부 언론 보도·기업DB "
         "표기이며 피치덱·재무제표는 없다 → 전 기업 '간이 진단' 대상.", False, 10),
        ("라우팅·밴드 판정", True, 11),
        ("트랙(500/HAX/바이오)과 스테이지 밴드는 공개 사실로 배정했다. 바이오 치료제·"
         "의약품 생산은 IndieBio 라우팅, 순수 SW는 500(HAX 제외 섹터), 물리 하드웨어는 "
         "HAX. 미확인 라운드는 보수적으로 낮은 밴드에 뒀다. 경계 사례는 게이트_라우팅 "
         "시트에 사유를 적었다.", False, 10),
        ("레벨 분류의 블라인드성", True, 11),
        ("각 기업의 4축 레벨은 dataset·정답을 본 적 없는 격리 세션이 개선된 §3·§4 규칙"
         "(판별 질문 포함)만 보고 분류했다. 회사의 조달 실적을 Team 근거로 쓰지 않는 등 "
         "증거 등급 규칙을 적용했다.", False, 10),
        ("이 표본으로 말할 수 있는 것 / 없는 것", True, 11),
        ("● 말할 수 있는 것: 엔진의 1차 필터(라우팅+게이트)가 스테이지·섹터를 실제로 "
         "거른다. 디캠프 배치는 500/HAX보다 뒤 단계라 다수가 게이트에서 탈락하는 것이 "
         "구조적으로 예측된다.", False, 10),
        ("● 말할 수 없는 것: 이들은 '500/HAX 지원·합격' 기업이 아니라 '디캠프 배치' "
         "선발사다. 따라서 이 판정은 '엔진이 500/HAX 심사를 재현하는가'의 증거가 아니라, "
         "'다른 프로그램 선발사를 500/HAX 기준으로 보면 어디서 걸리는가'의 매핑이다. "
         "정밀도·특이도는 이 표본으로 측정 불가.", False, 10),
    ]
    ws4.column_dimensions["A"].width = 110
    for i, (txt, bold, size) in enumerate(notes, 1):
        c = _cell(ws4, i, 1, txt, bold=bold, size=size, wrap=True)
        if bold and size >= 11:
            c.fill = SUB_FILL
        ws4.row_dimensions[i].height = 30 if len(txt) > 60 else 16

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(build())
