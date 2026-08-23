"""US FORGED — 산출물 워크북(§6) + 자체 채점·리젝트 감사(§8).

시트: 1) 요약+provenance+자체채점  2) 발송_후보(이메일 보유)  3) 연락처_확보_필요
      4) 검토(unclear)  5) 리젝트_감사(무작위 30, §8-⑤)  6) 중복_엔티티(§2)

출력 원칙(§0): "선발/요건충족" 표현 금지. 모든 통과행에 unverifiable_requirements 병기.
"""
from __future__ import annotations

import random
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from screening import uf_dedup, uf_shortlist, uf_snapshot

OUT = Path(__file__).resolve().parent.parent / "output" / "screening" / "us_forged_shortlist.xlsx"
FONT = "Arial"
HDR = PatternFill("solid", fgColor="1F3864")
HF = Font(name=FONT, bold=True, color="FFFFFF", size=9)
GREEN = PatternFill("solid", fgColor="E2EFDA")
YEL = PatternFill("solid", fgColor="FFF2CC")
GRY = PatternFill("solid", fgColor="EDEDED")
THIN = Side(style="thin", color="BFBFBF")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _c(ws, r, c, v, *, bold=False, fill=None, wrap=False, align="left", size=9):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=FONT, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORD
    if fill:
        cell.fill = fill
    return cell


def _hdr(ws, r, hs, ws_):
    from openpyxl.utils import get_column_letter
    for i, (h, w) in enumerate(zip(hs, ws_), 1):
        c = _c(ws, r, i, h, bold=True, fill=HDR, align="center"); c.font = HF
        ws.column_dimensions[get_column_letter(i)].width = w


def _contact_status(e):
    if e.get("email"):
        return "email"
    if e.get("website"):
        return "website_only"
    return "none"


def build(run_ts: str | None = None) -> Path:
    rows = uf_snapshot.load_rows()
    assessed = uf_shortlist.build(rows)
    prov = uf_snapshot.run_metadata(uf_snapshot.DEFAULT_SNAPSHOT,
                                    run_ts or datetime.now(timezone.utc).isoformat(timespec="seconds"),
                                    rows)
    disp = uf_shortlist.summarize(assessed)
    outreach = [a for a in assessed if a["disposition"] == "outreach"]
    review = [a for a in assessed if a["disposition"] == "review"]
    with_email = [a for a in outreach if a.get("email")]
    no_email = [a for a in outreach if not a.get("email")]

    wb = Workbook()
    # ---- 시트1 요약 + provenance + 자체채점(§8) ----
    ws = wb.active; ws.title = "요약_채점"; ws.sheet_view.showGridLines = False
    _c(ws, 1, 1, "US FORGED 발송 후보 선별 — 요약·provenance·자체채점", bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    r = 3
    _c(ws, r, 1, "provenance", bold=True, fill=GRY); r += 1
    for k in ("input_snapshot", "input_sha256", "input_rows", "engine_commit", "run_timestamp"):
        _c(ws, r, 1, k); _c(ws, r, 2, str(prov.get(k))); r += 1
    r += 1
    _c(ws, r, 1, "판정 분포(전 엔티티 %d)" % len(assessed), bold=True, fill=GRY); r += 1
    for k, v in disp.items():
        _c(ws, r, 1, k); _c(ws, r, 2, v, align="center"); r += 1
    r += 1
    # 자체채점 지표(§8)
    from screening import uf_golden
    _, gsum = uf_golden.evaluate_all()
    mp = gsum.get("classification_must_pass", {}); mf = gsum.get("classification_must_fail", {})
    ELEVEN = ["Robotics/Automation", "Advanced Manufacturing", "Energy/Climate Tech",
              "Industrial Hardware", "Semiconductor/Advanced Materials", "Sensor/Edge Device",
              "Physical AI", "Healthtech Device", "Manufacturing Process Innovation",
              "Aerospace", "Quantum"]
    ff = Counter(a["cls_matched_program_field"] for a in outreach)
    zero = [f for f in ELEVEN if ff.get(f, 0) == 0]
    stage_leak = sum(1 for a in assessed if a["disposition"] == "outreach"
                     and a["stage_bucket"] == uf_shortlist.uf_stage.OUT_OF_SCOPE)
    metrics = [
        ("골든 must_pass 통과율", f"{mp.get('pass')}/{mp.get('total')}"),
        ("골든 must_fail 차단율", f"{mf.get('pass')}/{mf.get('total')}"),
        ("스테이지 이탈 잔류(0이어야)", stage_leak),
        ("11개 분야 중 통과 0", ", ".join(zero) or "없음"),
        ("최종 발송 후보", len(outreach)),
        ("  이메일 보유 / 연락처 필요", f"{len(with_email)} / {len(no_email)}"),
        ("검토(unclear)", len(review)),
    ]
    _c(ws, r, 1, "자체 채점(§8)", bold=True, fill=GRY); r += 1
    for k, v in metrics:
        _c(ws, r, 1, k); _c(ws, r, 2, str(v), align="center"); r += 1

    # ---- 시트2/3 발송 후보 + 연락처 필요 ----
    def _write_candidates(sheet, data, title):
        w = wb.create_sheet(sheet); w.sheet_view.showGridLines = False
        _c(w, 1, 1, title, bold=True, size=12)
        w.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        _hdr(w, 3, ["국문명", "분야", "스테이지", "타겟시장", "confidence",
                    "판정근거(evidence)", "검증불가 요건", "연락처", "consumer_facing",
                    "maturity_signal", "merged_from"],
             [18, 22, 11, 12, 9, 40, 30, 10, 12, 22, 16])
        rr = 4
        for a in sorted(data, key=lambda x: (x["cls_matched_program_field"], x["name_ko"])):
            f = GREEN
            _c(w, rr, 1, a["name_ko"], fill=f)
            _c(w, rr, 2, a["cls_matched_program_field"], fill=f)
            _c(w, rr, 3, a.get("stage", ""), fill=f, align="center")
            _c(w, rr, 4, a.get("target") or "미상", fill=f, align="center", size=8)
            _c(w, rr, 5, a.get("cls_confidence", ""), fill=f, align="center", size=8)
            _c(w, rr, 6, (a.get("cls_evidence") or "")[:60], fill=f, size=8)
            _c(w, rr, 7, "; ".join(a.get("unverifiable_requirements", [])), fill=f, size=8)
            _c(w, rr, 8, _contact_status(a), fill=f, align="center", size=8)
            _c(w, rr, 9, "⚠검토" if a.get("cls_consumer_facing_end_product") else "", fill=f, align="center", size=8)
            _c(w, rr, 10, (a.get("cls_maturity_signal") or "")[:24], fill=f, size=8)
            _c(w, rr, 11, ",".join(a.get("merged_from", []) or []), fill=f, size=8)
            rr += 1
        w.freeze_panes = "A4"

    _write_candidates("발송_후보", with_email,
                      f"발송 후보(이메일 보유) {len(with_email)}개사 — 요건충족 아님, 설문 발송 우선순위")
    _write_candidates("연락처_확보_필요", no_email,
                      f"발송 후보(이메일 결측) {len(no_email)}개사 — Website 확보 후 발송")

    # ---- 시트4 검토(unclear) ----
    wr = wb.create_sheet("검토_unclear"); wr.sheet_view.showGridLines = False
    _c(wr, 1, 1, f"검토 대상(분류 저신뢰) {len(review)}개사 — 소개문만으론 판단 보류, 사람 확인",
       bold=True, size=12)
    wr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    _hdr(wr, 3, ["국문명", "스테이지", "판정근거(evidence)", "업종(CB)"], [20, 11, 50, 24])
    rr = 4
    for a in sorted(review, key=lambda x: x["name_ko"]):
        _c(wr, rr, 1, a["name_ko"]); _c(wr, rr, 2, a.get("stage", ""), align="center")
        _c(wr, rr, 3, (a.get("cls_evidence") or "")[:70], size=8)
        _c(wr, rr, 4, (a.get("industry") or "")[:24], size=8)
        rr += 1
    wr.freeze_panes = "A4"

    # ---- 시트5 리젝트 감사(§8-⑤): 무작위 30(sw15/consumer10/notstartup5) ----
    rnd = random.Random(20260821)
    pools = {"software_only": [a for a in assessed if a.get("cls_verdict") == "software_only"],
             "consumer": [a for a in assessed if a.get("cls_verdict") == "consumer"],
             "not_a_startup": [a for a in assessed if a.get("cls_verdict") == "not_a_startup"]}
    picks = []
    for verd, n in (("software_only", 15), ("consumer", 10), ("not_a_startup", 5)):
        lst = pools[verd][:]; rnd.shuffle(lst); picks += lst[:n]
    wa = wb.create_sheet("리젝트_감사"); wa.sheet_view.showGridLines = False
    _c(wa, 1, 1, "리젝트 감사 — 탈락 풀 무작위 30(sw15/consumer10/비스타트업5, 시드고정). "
       "재현율 확인용: 잘못 탈락한 게 없는지 사람이 읽는다(§8-⑤).", bold=True, size=11, wrap=True)
    wa.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    wa.row_dimensions[1].height = 30
    _hdr(wa, 3, ["verdict", "국문명", "1줄 소개(desc)", "판정근거(evidence)", "업종(CB)"],
         [14, 18, 48, 40, 20])
    rr = 4
    for a in picks:
        _c(wa, rr, 1, a.get("cls_verdict"), align="center", size=8)
        _c(wa, rr, 2, a["name_ko"], size=8)
        _c(wa, rr, 3, (a.get("desc") or "")[:70], size=8)
        _c(wa, rr, 4, (a.get("cls_evidence") or "")[:55], size=8)
        _c(wa, rr, 5, (a.get("industry") or "")[:20], size=8)
        rr += 1
    wa.freeze_panes = "A4"

    # ---- 시트6 중복 엔티티(§2) ----
    dups = uf_dedup.duplicate_report(rows)
    wd = wb.create_sheet("중복_엔티티"); wd.sheet_view.showGridLines = False
    _c(wd, 1, 1, f"동명 다중행 {len(dups)}건 — 신원 판정(§2). DB 관리자 정정 참고",
       bold=True, size=12)
    wd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _hdr(wd, 3, ["사명", "행수", "엔티티수", "신원", "플래그", "사업자번호들"],
         [20, 8, 10, 16, 24, 40])
    rr = 4
    for d in dups:
        _c(wd, rr, 1, d["name"]); _c(wd, rr, 2, d["rows"], align="center")
        _c(wd, rr, 3, d["entities"], align="center"); _c(wd, rr, 4, d["identity"], size=8)
        _c(wd, rr, 5, d["flags"], size=8); _c(wd, rr, 6, ", ".join(d["biz_nos"]), size=8)
        rr += 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT
